# reports/template_report_fixed.py - ФИКСИРОВАННЫЙ ГЕНЕРАТОР ОТЧЕТОВ ПО ШАБЛОНУ (ИСПРАВЛЕННАЯ ВЕРСИЯ)
import os
import shutil
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from typing import Dict, List, Any
import traceback

# Импортируем модели напрямую
from database.models import (
    Company, 
    UploadedFile, 
    Sheet1Structure, 
    Sheet2Demand, 
    Sheet3Balance, 
    Sheet4Supply, 
    Sheet5Sales,
    Sheet6Aviation,
    Sheet7Comments
)

class FixedTemplateReportGenerator:
    """Исправленный генератор отчетов по шаблону"""
    
    def __init__(self, db_connection, template_path: str = None):
        self.db = db_connection
        
        if template_path is None:
            self.template_path = os.path.join('report_templates', 'Сводный_отчет_шаблон.xlsx')
        else:
            self.template_path = template_path
        
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Шаблон не найден: {self.template_path}")
        
        print(f"Используем шаблон: {self.template_path}")
    
    def generate_report(self, report_date: datetime = None) -> str:
        """Генерация отчета по шаблону"""
        try:
            if report_date is None:
                report_date = datetime.now().date()
            
            print(f"\n=== ГЕНЕРАЦИЯ ОТЧЕТА ПО ШАБЛОНУ ===")
            print(f"Дата отчета: {report_date}")
            
            # Получаем данные из базы
            session = self.db.db.get_session()
            
            # Собираем данные для всех компаний
            companies_data = self._collect_company_data(session)
            
            session.close()
            
            if not companies_data:
                raise Exception("Нет данных в базе для генерации отчета")
            
            # Создаем копию шаблона
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f'Сводный_отчет_{timestamp}.xlsx'
            output_path = os.path.join('reports_output', output_filename)
            
            # Создаем папку если нет
            os.makedirs('reports_output', exist_ok=True)
            
            # Копируем шаблон
            shutil.copy2(self.template_path, output_path)
            
            # Открываем для заполнения
            wb = load_workbook(output_path)
            
            print(f"Листы в шаблоне: {wb.sheetnames}")
            
            # Заполняем данные
            self._fill_template(wb, companies_data, report_date)
            
            # Сохраняем
            wb.save(output_path)
            
            print(f"\n✅ Отчет создан: {output_path}")
            print(f"Компаний обработано: {len(companies_data)}")
            
            return output_path
            
        except Exception as e:
            print(f"❌ Ошибка генерации отчета: {e}")
            traceback.print_exc()
            raise
    
    def _collect_company_data(self, session) -> Dict[str, Dict[str, Any]]:
        """Сбор данных по компаниям из базы"""
        companies_data = {}
        
        print("Сбор данных из базы...")
        
        # 1. Собираем остатки (Sheet3Balance)
        balances = session.query(Sheet3Balance).all()
        print(f"Найдено записей остатков: {len(balances)}")
        
        for balance in balances:
            company_name = balance.company_name
            if company_name not in companies_data:
                companies_data[company_name] = {
                    'name': company_name,
                    'stock_ai92': 0,
                    'stock_ai95': 0,
                    'stock_diesel_winter': 0,
                    'stock_diesel_arctic': 0,
                    'sales_ai92': 0,
                    'sales_ai95': 0,
                    'supply_ai92': 0,
                    'supply_ai95': 0,
                    'demand_ai92': 0,
                    'demand_ai95': 0,
                }
            
            companies_data[company_name]['stock_ai92'] += (balance.stock_ai92 or 0)
            companies_data[company_name]['stock_ai95'] += (balance.stock_ai95 or 0)
            companies_data[company_name]['stock_diesel_winter'] += (balance.stock_diesel_winter or 0)
            companies_data[company_name]['stock_diesel_arctic'] += (balance.stock_diesel_arctic or 0)
        
        # 2. Собираем реализацию (Sheet5Sales)
        sales = session.query(Sheet5Sales).all()
        print(f"Найдено записей реализации: {len(sales)}")
        
        for sale in sales:
            company_name = sale.company_name
            if company_name not in companies_data:
                companies_data[company_name] = {
                    'name': company_name,
                    'stock_ai92': 0,
                    'stock_ai95': 0,
                    'stock_diesel_winter': 0,
                    'stock_diesel_arctic': 0,
                    'sales_ai92': 0,
                    'sales_ai95': 0,
                    'supply_ai92': 0,
                    'supply_ai95': 0,
                    'demand_ai92': 0,
                    'demand_ai95': 0,
                }
            
            companies_data[company_name]['sales_ai92'] += (sale.monthly_ai92 or 0)
            companies_data[company_name]['sales_ai95'] += (sale.monthly_ai95 or 0)
        
        # 3. Собираем поставки (Sheet4Supply)
        supplies = session.query(Sheet4Supply).all()
        print(f"Найдено записей поставок: {len(supplies)}")
        
        for supply in supplies:
            company_name = supply.company_name
            if company_name not in companies_data:
                companies_data[company_name] = {
                    'name': company_name,
                    'stock_ai92': 0,
                    'stock_ai95': 0,
                    'stock_diesel_winter': 0,
                    'stock_diesel_arctic': 0,
                    'sales_ai92': 0,
                    'sales_ai95': 0,
                    'supply_ai92': 0,
                    'supply_ai95': 0,
                    'demand_ai92': 0,
                    'demand_ai95': 0,
                }
            
            companies_data[company_name]['supply_ai92'] += (supply.supply_ai92 or 0)
            companies_data[company_name]['supply_ai95'] += (supply.supply_ai95 or 0)
        
        # 4. Собираем потребность (Sheet2Demand)
        demands = session.query(Sheet2Demand).all()
        print(f"Найдено записей потребности: {len(demands)}")
        
        for demand in demands:
            # Используем связанную компанию или ищем по ID
            if demand.company_id:
                company = session.query(Company).get(demand.company_id)
                if company:
                    company_name = company.name
                else:
                    company_name = "Неизвестная компания"
            else:
                company_name = "Неизвестная компания"
            
            if company_name not in companies_data:
                companies_data[company_name] = {
                    'name': company_name,
                    'stock_ai92': 0,
                    'stock_ai95': 0,
                    'stock_diesel_winter': 0,
                    'stock_diesel_arctic': 0,
                    'sales_ai92': 0,
                    'sales_ai95': 0,
                    'supply_ai92': 0,
                    'supply_ai95': 0,
                    'demand_ai92': 0,
                    'demand_ai95': 0,
                }
            
            companies_data[company_name]['demand_ai92'] += (demand.gasoline_ai92 or 0)
            companies_data[company_name]['demand_ai95'] += (demand.gasoline_ai95 or 0)
        
        print(f"Собраны данные для {len(companies_data)} компаний:")
        for name, data in companies_data.items():
            print(f"  - {name}: AI92={data['stock_ai92']:.3f}, AI95={data['stock_ai95']:.3f}")
        
        return companies_data
    
    def _fill_template(self, wb, companies_data: Dict[str, Dict[str, Any]], report_date):
        """Заполнение шаблона данными"""
        print("Заполнение шаблона...")
        
        # Создаем список компаний для поиска
        company_names = list(companies_data.keys())
        
        # 1. Заполняем лист "2-Потребность" если он есть
        if '2-Потребность' in wb.sheetnames:
            print("  Заполнение листа '2-Потребность'...")
            self._fill_sheet_with_data(wb['2-Потребность'], companies_data, company_names, 'demand')
        
        # 2. Заполняем лист "3-Остатки" если он есть
        if '3-Остатки' in wb.sheetnames:
            print("  Заполнение листа '3-Остатки'...")
            self._fill_sheet_with_data(wb['3-Остатки'], companies_data, company_names, 'stock')
        
        # 3. Заполняем лист "4-Поставка" если он есть
        if '4-Поставка' in wb.sheetnames:
            print("  Заполнение листа '4-Поставка'...")
            self._fill_sheet_with_data(wb['4-Поставка'], companies_data, company_names, 'supply')
        
        # 4. Заполняем лист "5-Реализация" если он есть
        if '5-Реализация' in wb.sheetnames:
            print("  Заполнение листа '5-Реализация'...")
            self._fill_sheet_with_data(wb['5-Реализация'], companies_data, company_names, 'sales')
        
        # 5. Обновляем дату отчета
        self._update_report_date(wb, report_date)
    
    def _fill_sheet_with_data(self, ws, companies_data, company_names, data_type):
        """Универсальный метод заполнения листа"""
        # Ищем все строки с компаниями
        for row in range(1, 200):  # Проверяем больше строк
            for col in range(1, 15):  # Проверяем больше колонок
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    cell_value = str(cell.value).strip()
                    
                    # Ищем совпадение с названием компании
                    for company_name in company_names:
                        # Проверяем частичное совпадение
                        company_words = company_name.lower().split()
                        if any(word in cell_value.lower() for word in company_words if len(word) > 3):
                            # Нашли компанию, заполняем данные
                            data = companies_data.get(company_name)
                            if data:
                                self._fill_company_row(ws, row, data, data_type, company_name)
                            break
    
    def _fill_company_row(self, ws, row, data, data_type, company_name):
        """Заполнение строки компании"""
        if data_type == 'demand':
            # Заполняем потребность
            for col in range(1, 20):
                header_cell = ws.cell(row=row-1, column=col) if row > 1 else None
                if header_cell and header_cell.value:
                    header = str(header_cell.value).lower()
                    if '92' in header and any(x in header for x in ['аи', 'ai', 'бензин']):
                        ws.cell(row=row, column=col, value=data['demand_ai92'] or 0)
                        print(f"    {company_name}: потребность AI-92 = {data['demand_ai92'] or 0}")
                    elif '95' in header and any(x in header for x in ['аи', 'ai', 'бензин']):
                        ws.cell(row=row, column=col, value=data['demand_ai95'] or 0)
                        print(f"    {company_name}: потребность AI-95 = {data['demand_ai95'] or 0}")
        
        elif data_type == 'stock':
            # Заполняем остатки
            for col in range(1, 20):
                header_cell = ws.cell(row=row-1, column=col) if row > 1 else None
                if header_cell and header_cell.value:
                    header = str(header_cell.value).lower()
                    if '92' in header and any(x in header for x in ['аи', 'ai', 'остат']):
                        ws.cell(row=row, column=col, value=data['stock_ai92'] or 0)
                        print(f"    {company_name}: остатки AI-92 = {data['stock_ai92'] or 0}")
                    elif '95' in header and any(x in header for x in ['аи', 'ai', 'остат']):
                        ws.cell(row=row, column=col, value=data['stock_ai95'] or 0)
                        print(f"    {company_name}: остатки AI-95 = {data['stock_ai95'] or 0}")
                    elif 'зим' in header and any(x in header for x in ['дизель', 'диз', 'остат']):
                        ws.cell(row=row, column=col, value=data['stock_diesel_winter'] or 0)
                        print(f"    {company_name}: дизель зимний = {data['stock_diesel_winter'] or 0}")
        
        elif data_type == 'supply':
            # Заполняем поставки
            for col in range(1, 20):
                header_cell = ws.cell(row=row-1, column=col) if row > 1 else None
                if header_cell and header_cell.value:
                    header = str(header_cell.value).lower()
                    if '92' in header and any(x in header for x in ['аи', 'ai', 'постав']):
                        ws.cell(row=row, column=col, value=data['supply_ai92'] or 0)
                        print(f"    {company_name}: поставки AI-92 = {data['supply_ai92'] or 0}")
                    elif '95' in header and any(x in header for x in ['аи', 'ai', 'постав']):
                        ws.cell(row=row, column=col, value=data['supply_ai95'] or 0)
                        print(f"    {company_name}: поставки AI-95 = {data['supply_ai95'] or 0}")
        
        elif data_type == 'sales':
            # Заполняем реализацию
            for col in range(1, 20):
                header_cell = ws.cell(row=row-1, column=col) if row > 1 else None
                if header_cell and header_cell.value:
                    header = str(header_cell.value).lower()
                    if '92' in header and any(x in header for x in ['аи', 'ai', 'реализ', 'продаж']):
                        ws.cell(row=row, column=col, value=data['sales_ai92'] or 0)
                        print(f"    {company_name}: реализация AI-92 = {data['sales_ai92'] or 0}")
                    elif '95' in header and any(x in header for x in ['аи', 'ai', 'реализ', 'продаж']):
                        ws.cell(row=row, column=col, value=data['sales_ai95'] or 0)
                        print(f"    {company_name}: реализация AI-95 = {data['sales_ai95'] or 0}")
    
    def _update_report_date(self, wb, report_date):
        """Обновление даты отчета"""
        date_str = report_date.strftime('%d.%m.%Y')
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in range(1, 20):
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        if any(x in cell.value.lower() for x in ['дата', 'отчет', 'состоян', 'date']):
                            # Пробуем заполнить соседнюю ячейку
                            if col < 20:
                                date_cell = ws.cell(row=row, column=col+1)
                                date_cell.value = date_str
                                print(f"    Обновлена дата в {sheet_name}: {date_cell.coordinate} = {date_str}")