# reports/template_manager.py
import os
import shutil
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class TemplateManager:
    def __init__(self, templates_dir='report_templates'):
        self.templates_dir = templates_dir
        os.makedirs(templates_dir, exist_ok=True)
    
    def create_sample_template(self):
        """Создание образца шаблона отчета"""
        template_path = os.path.join(self.templates_dir, 'sample_template.xlsx')
        
        wb = openpyxl.Workbook()
        
        # Лист 1: Титульная страница
        ws1 = wb.active
        ws1.title = "Титульная страница"
        
        # Заголовок
        ws1['A1'] = "СВОДНЫЙ ОТЧЕТ ПО ТОПЛИВООБЕСПЕЧЕНИЮ"
        ws1['A1'].font = Font(size=16, bold=True)
        ws1.merge_cells('A1:H1')
        
        ws1['A2'] = "Дата отчета: {report_date}"
        ws1['A2'].font = Font(bold=True)
        ws1.merge_cells('A2:H2')
        
        ws1['A3'] = "Период: {period}"
        ws1.merge_cells('A3:H3')
        
        ws1['A5'] = "Данные автоматически сформированы системой"
        ws1['A5'].font = Font(italic=True, color="666666")
        
        # Лист 2: Сводная таблица
        ws2 = wb.create_sheet(title="Сводная таблица")
        
        headers = ["№", "Компания", "Кол-во АЗС", "Работающих АЗС", 
                  "Потребность бензин", "Потребность дизель",
                  "Остатки АИ-92", "Остатки АИ-95", "Реализация АИ-92"]
        
        row = 1
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Пример данных с заполнителями
        companies = ["{company1}", "{company2}", "{company3}"]
        for i, company in enumerate(companies, 1):
            ws2.cell(row=i+1, column=1, value=i)
            ws2.cell(row=i+1, column=2, value=company)
            # Остальные ячейки будут заполняться данными
        
        # Лист 3: Графики (пустой для примера)
        ws3 = wb.create_sheet(title="Графики")
        ws3['A1'] = "Графики и диаграммы будут сгенерированы автоматически"
        
        wb.save(template_path)
        return template_path
    
    def fill_template(self, template_path, data, output_path):
        """Заполнение шаблона данными"""
        wb = openpyxl.load_workbook(template_path)
        
        # Заполняем титульную страницу
        if "Титульная страница" in wb.sheetnames:
            ws = wb["Титульная страница"]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Заменяем заполнители
                        cell.value = cell.value.replace('{report_date}', data.get('report_date', ''))
                        cell.value = cell.value.replace('{period}', data.get('period', ''))
        
        # Заполняем сводную таблицу
        if "Сводная таблица" in wb.sheetnames:
            ws = wb["Сводная таблица"]
            row = 2  # Начинаем со второй строки (после заголовков)
            
            for company_data in data.get('companies', []):
                ws.cell(row=row, column=1, value=row-1)  
                ws.cell(row=row, column=2, value=company_data.get('name', ''))
                ws.cell(row=row, column=3, value=company_data.get('azs_count', 0))
                ws.cell(row=row, column=4, value=company_data.get('working_azs', 0))
                ws.cell(row=row, column=5, value=company_data.get('gasoline_demand', 0))
                ws.cell(row=row, column=6, value=company_data.get('diesel_demand', 0))
                ws.cell(row=row, column=7, value=company_data.get('stock_ai92', 0))
                ws.cell(row=row, column=8, value=company_data.get('stock_ai95', 0))
                ws.cell(row=row, column=9, value=company_data.get('sales_ai92', 0))
                row += 1
        
        wb.save(output_path)
        return output_path