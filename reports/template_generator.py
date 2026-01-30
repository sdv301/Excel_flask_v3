# reports/template_generator.py - АДАПТИРОВАННЫЙ ПОД ВАШ ШАБЛОН
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from typing import Dict, List, Any, Optional
import traceback

class TemplateFiller:
    def __init__(self, template_path: str = None):
        if template_path is None:
            self.template_path = os.path.join('report_templates', 'Сводный_отчет_шаблон.xlsx')
        else:
            self.template_path = template_path
        
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Шаблон не найден: {self.template_path}")
    
    def fill_template(self, data: Dict[str, Any], output_path: str = None) -> str:
        """Заполнение шаблона данными"""
        try:
            print(f"Загрузка шаблона: {self.template_path}")
            
            wb = load_workbook(self.template_path)
            
            print(f"Листы в шаблоне: {wb.sheetnames}")
            
            # Проверяем наличие нужных листов
            available_sheets = set(wb.sheetnames)
            required_sheets = {'2-Потребность', '3-Остатки', '4-Поставка', '5-Реализация'}
            
            missing_sheets = required_sheets - available_sheets
            if missing_sheets:
                print(f"ВНИМАНИЕ: Отсутствуют листы: {missing_sheets}")
                print(f"Доступные листы: {available_sheets}")
            
            # Заполняем только те листы, которые есть в шаблоне
            if '2-Потребность' in wb.sheetnames:
                print("\n=== ЗАПОЛНЕНИЕ ЛИСТА 2: ПОТРЕБНОСТЬ ===")
                self._fill_sheet_demand(wb['2-Потребность'], data)
            
            if '3-Остатки' in wb.sheetnames:
                print("\n=== ЗАПОЛНЕНИЕ ЛИСТА 3: ОСТАТКИ ===")
                self._fill_sheet_balance(wb['3-Остатки'], data)
            
            if '4-Поставка' in wb.sheetnames:
                print("\n=== ЗАПОЛНЕНИЕ ЛИСТА 4: ПОСТАВКИ ===")
                self._fill_sheet_supply(wb['4-Поставка'], data)
            
            if '5-Реализация' in wb.sheetnames:
                print("\n=== ЗАПОЛНЕНИЕ ЛИСТА 5: РЕАЛИЗАЦИЯ ===")
                self._fill_sheet_sales(wb['5-Реализация'], data)
            
            # Обновляем титульный лист если есть
            if '1-Структура' in wb.sheetnames:
                self._update_summary_sheet(wb['1-Структура'], data)
            
            # Сохраняем файл
            if output_path is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_dir = 'reports_output'
                os.makedirs(output_dir, exist_ok=True)
                output_path = os.path.join(output_dir, f'Сводный_отчет_{timestamp}.xlsx')
            
            wb.save(output_path)
            print(f"\n✅ Шаблон заполнен и сохранен: {output_path}")
            
            return output_path
            
        except Exception as e:
            print(f"❌ Ошибка заполнения шаблона: {e}")
            traceback.print_exc()
            raise
    
    def _update_summary_sheet(self, ws, data: Dict[str, Any]):
        """Обновление сводного листа (1-Структура)"""
        try:
            report_date = data.get('report_date', datetime.now())
            if isinstance(report_date, datetime):
                report_date_str = report_date.strftime('%d.%m.%Y')
            else:
                report_date_str = str(report_date)
            
            companies = data.get('companies', [])
            total_companies = len(companies)
            total_azs = sum(c.get('azs_count', 0) for c in companies)
            
            print(f"Обновление сводного листа: {total_companies} компаний, {total_azs} АЗС")
            
            # Ищем ячейки для заполнения
            for row in range(1, 30):
                for col in range(1, 10):
                    try:
                        cell = ws.cell(row=row, column=col)
                        if cell.value and isinstance(cell.value, str):
                            cell_str = str(cell.value)
                            if '{дата_отчета}' in cell_str:
                                cell.value = cell_str.replace('{дата_отчета}', report_date_str)
                            elif '{кол_во_компаний}' in cell_str:
                                cell.value = cell_str.replace('{кол_во_компаний}', str(total_companies))
                            elif '{всего_азс}' in cell_str:
                                cell.value = cell_str.replace('{всего_азс}', str(total_azs))
                    except:
                        continue
        except Exception as e:
            print(f"Ошибка обновления сводного листа: {e}")
    
    def _fill_sheet_demand(self, ws, data: Dict[str, Any]):
        """Заполнение листа 2-Потребность"""
        companies = data.get('companies', [])
        print(f"Компаний для заполнения потребности: {len(companies)}")
        
        # Ищем таблицу потребности
        table_start = self._find_table_start(ws, search_text=['ГОД', 'МЕСЯЦ', 'Бензин', 'Дизель'])
        
        if table_start is None:
            print("Не найдена таблица потребности")
            return
        
        print(f"Таблица потребности начинается со строки: {table_start}")
        
        # Заполняем данные
        row = table_start
        for i, company in enumerate(companies, 1):
            if row > table_start + 50:  # Ограничение
                break
            
            print(f"  Компания {i}: {company.get('name', '')}")
            
            # Ищем безопасные ячейки для записи
            for col in range(1, 15):
                try:
                    cell = ws.cell(row=row, column=col)
                    # Пропускаем объединенные ячейки
                    if hasattr(cell, 'is_merged') and cell.is_merged:
                        continue
                    
                    # Заполняем данные в зависимости от структуры таблицы
                    if col == 1:  # Номер
                        cell.value = i
                    elif col == 2:  # Название компании
                        cell.value = company.get('name', '')
                    elif col == 3:  # Бензин всего (год)
                        cell.value = company.get('gasoline_demand', 0)
                        cell.number_format = '0.00'
                    elif col == 4:  # Дизель всего (год)
                        cell.value = company.get('diesel_demand', 0)
                        cell.number_format = '0.00'
                    elif col == 5:  # Бензин (месяц)
                        cell.value = company.get('monthly_gasoline', 0)
                        cell.number_format = '0.00'
                    elif col == 6:  # Дизель (месяц)
                        cell.value = company.get('monthly_diesel', 0)
                        cell.number_format = '0.00'
                    
                except Exception as e:
                    continue
            
            row += 1
        
        # Добавляем итоговую строку
        if row > table_start and row < table_start + 100:
            self._add_simple_total_row(ws, table_start, row - 1, row, [3, 4, 5, 6])
    
    def _fill_sheet_balance(self, ws, data: Dict[str, Any]):
        """Заполнение листа 3-Остатки"""
        companies = data.get('companies', [])
        print(f"Компаний для заполнения остатков: {len(companies)}")
        
        # Ищем таблицу остатков
        table_start = self._find_table_start(ws, search_text=['Остатки', 'АИ-92', 'АИ-95', 'Дизель'])
        
        if table_start is None:
            print("Не найдена таблица остатков")
            return
        
        print(f"Таблица остатков начинается со строки: {table_start}")
        
        row = table_start
        for i, company in enumerate(companies, 1):
            if row > table_start + 50:
                break
            
            print(f"  Компания {i}: {company.get('name', '')}")
            
            for col in range(1, 15):
                try:
                    cell = ws.cell(row=row, column=col)
                    if hasattr(cell, 'is_merged') and cell.is_merged:
                        continue
                    
                    if col == 1:  # Номер
                        cell.value = i
                    elif col == 2:  # Компания
                        cell.value = company.get('name', '')
                    elif col == 3:  # АИ-92
                        cell.value = company.get('stock_ai92', 0)
                        cell.number_format = '0.00'
                    elif col == 4:  # АИ-95
                        cell.value = company.get('stock_ai95', 0)
                        cell.number_format = '0.00'
                    elif col == 5:  # Дизель зимний
                        cell.value = company.get('stock_diesel_winter', 0)
                        cell.number_format = '0.00'
                    elif col == 6:  # Дизель арктический
                        cell.value = company.get('stock_diesel_arctic', 0)
                        cell.number_format = '0.00'
                    
                except:
                    continue
            
            row += 1
        
        # Итоговая строка
        if row > table_start and row < table_start + 100:
            self._add_simple_total_row(ws, table_start, row - 1, row, [3, 4, 5, 6])
    
    def _fill_sheet_supply(self, ws, data: Dict[str, Any]):
        """Заполнение листа 4-Поставка"""
        companies = data.get('companies', [])
        print(f"Компаний для заполнения поставок: {len(companies)}")
        
        # Ищем таблицу поставок
        table_start = self._find_table_start(ws, search_text=['Поставка', 'Срок', 'Объем', 'АИ-92'])
        
        if table_start is None:
            print("Не найдена таблица поставок")
            return
        
        print(f"Таблица поставок начинается со строки: {table_start}")
        
        row = table_start
        for i, company in enumerate(companies, 1):
            if row > table_start + 50:
                break
            
            print(f"  Компания {i}: {company.get('name', '')}")
            
            for col in range(1, 15):
                try:
                    cell = ws.cell(row=row, column=col)
                    if hasattr(cell, 'is_merged') and cell.is_merged:
                        continue
                    
                    if col == 1:  # Номер
                        cell.value = i
                    elif col == 2:  # Компания
                        cell.value = company.get('name', '')
                    elif col == 3:  # АИ-92 (поставки)
                        cell.value = company.get('supply_ai92', 0)
                        cell.number_format = '0.00'
                    elif col == 4:  # АИ-95 (поставки)
                        cell.value = company.get('supply_ai95', 0)
                        cell.number_format = '0.00'
                    elif col == 5:  # Дизель зимний (поставки)
                        cell.value = company.get('supply_diesel_winter', 0)
                        cell.number_format = '0.00'
                    elif col == 6:  # Дизель арктический (поставки)
                        cell.value = company.get('supply_diesel_arctic', 0)
                        cell.number_format = '0.00'
                    
                except:
                    continue
            
            row += 1
        
        # Итоговая строка
        if row > table_start and row < table_start + 100:
            self._add_simple_total_row(ws, table_start, row - 1, row, [3, 4, 5, 6])
    
    def _fill_sheet_sales(self, ws, data: Dict[str, Any]):
        """Заполнение листа 5-Реализация"""
        companies = data.get('companies', [])
        print(f"Компаний для заполнения реализации: {len(companies)}")
        
        # Ищем таблицу реализации
        table_start = self._find_table_start(ws, search_text=['Реализация', 'Продажи', 'АИ-92', 'АИ-95'])
        
        if table_start is None:
            print("Не найдена таблица реализации")
            return
        
        print(f"Таблица реализации начинается со строки: {table_start}")
        
        row = table_start
        for i, company in enumerate(companies, 1):
            if row > table_start + 50:
                break
            
            print(f"  Компания {i}: {company.get('name', '')}")
            
            for col in range(1, 15):
                try:
                    cell = ws.cell(row=row, column=col)
                    if hasattr(cell, 'is_merged') and cell.is_merged:
                        continue
                    
                    if col == 1:  # Номер
                        cell.value = i
                    elif col == 2:  # Компания
                        cell.value = company.get('name', '')
                    elif col == 3:  # АИ-92 (реализация)
                        cell.value = company.get('sales_ai92', 0)
                        cell.number_format = '0.00'
                    elif col == 4:  # АИ-95 (реализация)
                        cell.value = company.get('sales_ai95', 0)
                        cell.number_format = '0.00'
                    elif col == 5:  # Дизель зимний (реализация)
                        cell.value = company.get('sales_diesel_winter', 0)
                        cell.number_format = '0.00'
                    elif col == 6:  # Дизель арктический (реализация)
                        cell.value = company.get('sales_diesel_arctic', 0)
                        cell.number_format = '0.00'
                    
                except:
                    continue
            
            row += 1
        
        # Итоговая строка
        if row > table_start and row < table_start + 100:
            self._add_simple_total_row(ws, table_start, row - 1, row, [3, 4, 5, 6])
    
    def _find_table_start(self, ws, search_text: List[str]) -> Optional[int]:
        """Находит начало таблицы для заполнения"""
        # Ищем заголовок таблицы
        for row in range(1, 100):
            for col in range(1, 20):
                try:
                    cell = ws.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        cell_value = str(cell.value).lower()
                        for text in search_text:
                            if text.lower() in cell_value:
                                print(f"    Найден заголовок: '{cell.value}' в {cell.coordinate}")
                                
                                # Ищем начало данных (пустую строку после заголовка)
                                for next_row in range(row + 1, row + 20):
                                    test_cell = ws.cell(row=next_row, column=1)
                                    if test_cell.value is None or test_cell.value == '':
                                        return next_row
                                
                                return row + 1  # Запасной вариант
                except Exception as e:
                    continue
        
        print(f"    Не найден заголовок таблицы: {search_text}")
        return None
    
    def _add_simple_total_row(self, ws, start_row: int, end_row: int, total_row: int, sum_columns: List[int]):
        """Добавляет простую итоговую строку без сложных формул"""
        try:
            # Добавляем надпись "ИТОГО"
            for col in range(1, 5):
                try:
                    cell = ws.cell(row=total_row, column=col)
                    if not hasattr(cell, 'is_merged') or not cell.is_merged:
                        if col == 2:  # Обычно вторая колонка для названия
                            cell.value = "ИТОГО:"
                            cell.font = Font(bold=True)
                        break
                except:
                    continue
            
            # Для числовых колонок добавляем суммы
            for col in sum_columns:
                try:
                    cell = ws.cell(row=total_row, column=col)
                    if hasattr(cell, 'is_merged') and cell.is_merged:
                        continue
                    
                    # Просто вычисляем сумму и записываем значение
                    total = 0
                    for r in range(start_row, end_row + 1):
                        try:
                            value_cell = ws.cell(row=r, column=col)
                            if value_cell.value and isinstance(value_cell.value, (int, float)):
                                total += float(value_cell.value)
                        except:
                            continue
                    
                    cell.value = total
                    cell.font = Font(bold=True)
                    cell.number_format = '0.00'
                    cell.fill = PatternFill(
                        start_color="FFFF99", end_color="FFFF99", fill_type="solid"
                    )
                    
                    print(f"    Итого в колонке {col}: {total}")
                    
                except Exception as e:
                    print(f"    Ошибка в колонке {col}: {e}")
        except Exception as e:
            print(f"    Ошибка добавления итоговой строки: {e}")