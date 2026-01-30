# reports/report_generator.py
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta
import os
from typing import List, Dict, Any
import json
from reports.template_manager import TemplateManager
from reports.template_generator import TemplateFiller



class SummaryReportGenerator:
    def __init__(self, db_connection):
        self.db = db_connection
        self.template_manager = TemplateManager()
    
    def generate_from_template(self, template_name: str, report_date: datetime = None, output_path: str = None) -> str:
        """Генерация отчета по шаблону"""
        if report_date is None:
            report_date = datetime.now().date()
        
        # Получаем агрегированные данные
        aggregated_data = self.db.get_aggregated_data(report_date)
        
        if not aggregated_data:
            raise Exception("Нет данных для генерации отчета")
        
        print(f"Получены данные для {len(aggregated_data)} компаний:")
        for company_name in aggregated_data.keys():
            print(f"  - {company_name}")
        
        # Формируем данные для шаблона
        template_data = {
            'report_date': report_date.strftime('%d.%m.%Y'),
            'period': f"{report_date.strftime('%B %Y')}",
            'companies': []
        }
        
        # Преобразуем агрегированные данные в формат для шаблона
        for company_name, company_data in aggregated_data.items():
            if not isinstance(company_data, dict):
                continue
            
            # Собираем данные компании
            company_info = {
                'name': company_name,
                'azs_count': 0,
                'working_azs': 0,
                'gasoline_demand': 0,
                'diesel_demand': 0,
                'stock_ai92': 0,
                'stock_ai95': 0,
                'sales_ai92': 0,
                'sales_ai95': 0
            }
            
            # Лист 1: АЗС
            sheet1_data = company_data.get('sheet1', [])
            if isinstance(sheet1_data, list):
                for item in sheet1_data:
                    if isinstance(item, dict):
                        company_info['azs_count'] += item.get('azs_count', 0)
                        company_info['working_azs'] += item.get('working_azs_count', 0)
            
            # Лист 2: Потребность
            sheet2_data = company_data.get('sheet2', {})
            if isinstance(sheet2_data, dict):
                company_info['gasoline_demand'] = sheet2_data.get('gasoline_total', 0)
                company_info['diesel_demand'] = sheet2_data.get('diesel_total', 0)
            
            # Лист 3: Остатки
            sheet3_totals = company_data.get('sheet3_totals', {})
            if isinstance(sheet3_totals, dict):
                company_info['stock_ai92'] = sheet3_totals.get('total_stock_ai92', 0)
                company_info['stock_ai95'] = sheet3_totals.get('total_stock_ai95', 0)
            
            # Лист 5: Реализация
            sheet5_totals = company_data.get('sheet5_totals', {})
            if isinstance(sheet5_totals, dict):
                company_info['sales_ai92'] = sheet5_totals.get('total_monthly_ai92', 0)
                company_info['sales_ai95'] = sheet5_totals.get('total_monthly_ai95', 0)
            
            template_data['companies'].append(company_info)
        
        # Проверяем наличие шаблона
        template_path = os.path.join('report_templates', template_name)
        if not os.path.exists(template_path):
            # Создаем образец шаблона
            template_path = self.template_manager.create_sample_template()
        
        # Генерируем отчет
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = os.path.join('reports_output', f'Отчет_по_шаблону_{timestamp}.xlsx')
        
        # Заполняем шаблон
        result_path = self.template_manager.fill_template(template_path, template_data, output_path)
        
        return result_path
    
    def create_custom_report(self, report_date: datetime = None, 
                           companies: List[str] = None,
                           include_sheets: List[str] = None) -> str:
        """Создание кастомного отчета с выбором компаний и листов"""
        if report_date is None:
            report_date = datetime.now().date()
        
        # Получаем данные
        aggregated_data = self.db.get_aggregated_data(report_date)
        
        if not aggregated_data:
            raise Exception("Нет данных для генерации отчета")
        
        # Фильтруем компании если нужно
        if companies:
            filtered_data = {k: v for k, v in aggregated_data.items() if k in companies}
        else:
            filtered_data = aggregated_data
        
        # Создаем отчет
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Всегда добавляем сводный лист
        self._create_summary_sheet(wb, filtered_data, report_date)
        
        # Добавляем выбранные листы
        include_sheets = include_sheets or ['structure', 'demand', 'balance', 'sales']
        
        if 'structure' in include_sheets:
            self._create_structure_sheet(wb, filtered_data, report_date)
        
        if 'demand' in include_sheets:
            self._create_demand_sheet(wb, filtered_data, report_date)
        
        if 'supply' in include_sheets:
            self._create_supply_summary_sheet(wb, filtered_data, report_date)
        
        if 'balance' in include_sheets:
            self._create_balance_summary_sheet(wb, filtered_data, report_date)
        
        if 'sales' in include_sheets:
            self._create_sales_summary_sheet(wb, filtered_data, report_date)
        
        # Добавляем лист с итогами
        self._create_totals_sheet(wb, filtered_data, report_date)
        
        # Сохраняем
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = os.path.join('reports_output', f'Кастомный_отчет_{timestamp}.xlsx')
        
        wb.save(output_path)
        return output_path
    
    def _create_totals_sheet(self, wb, data, report_date):
        """Лист с итоговыми показателями"""
        ws = wb.create_sheet(title="Итоги")
        
        ws['A1'] = "ИТОГОВЫЕ ПОКАЗАТЕЛИ"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Рассчитываем итоги
        total_azs = 0
        total_gasoline = 0
        total_diesel = 0
        total_stock = 0
        total_sales = 0
        
        for company_data in data.values():
            if not isinstance(company_data, dict):
                continue
            
            # АЗС
            sheet1_data = company_data.get('sheet1', [])
            if isinstance(sheet1_data, list):
                for item in sheet1_data:
                    if isinstance(item, dict):
                        total_azs += item.get('azs_count', 0)
            
            # Потребность
            sheet2_data = company_data.get('sheet2', {})
            if isinstance(sheet2_data, dict):
                total_gasoline += sheet2_data.get('gasoline_total', 0)
                total_diesel += sheet2_data.get('diesel_total', 0)
            
            # Остатки
            sheet3_totals = company_data.get('sheet3_totals', {})
            if isinstance(sheet3_totals, dict):
                total_stock += sheet3_totals.get('total_stock_ai92', 0)
                total_stock += sheet3_totals.get('total_stock_ai95', 0)
            
            # Реализация
            sheet5_totals = company_data.get('sheet5_totals', {})
            if isinstance(sheet5_totals, dict):
                total_sales += sheet5_totals.get('total_monthly_ai92', 0)
        
        # Заполняем таблицу
        indicators = [
            ["Общее количество компаний", len(data)],
            ["Всего АЗС", total_azs],
            ["Потребность в бензине (т/год)", total_gasoline],
            ["Потребность в дизтопливе (т/год)", total_diesel],
            ["Общие остатки топлива (т)", total_stock],
            ["Реализация АИ-92 (т/мес)", total_sales]
        ]
        
        row = 3
        for indicator, value in indicators:
            ws.cell(row=row, column=1, value=indicator)
            ws.cell(row=row, column=2, value=value)
            if isinstance(value, (int, float)) and value != len(data):
                ws.cell(row=row, column=2).number_format = '0.00'
            row += 1
        
        self._apply_default_styles(ws)
        self._auto_adjust_columns(ws)
    
    def _create_companies_sheet(self, wb, data, report_date):
        """Лист с детализацией по компаниям"""
        ws = wb.create_sheet(title="Компании")
        
        ws['A1'] = "ДЕТАЛИЗАЦИЯ ПО КОМПАНИЯМ"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        row = 3
        
        for company_name, company_data in data.items():
            if not isinstance(company_data, dict):
                continue
                
            ws.cell(row=row, column=1, value=f"Компания: {company_name}")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # Заголовки таблицы
            headers = ["Наименование", "Нефтебаз", "АЗС", "Работающих АЗС", "Принадлежность"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            
            row += 1
            
            # Данные - проверяем и нормализуем
            sheet1_data = company_data.get('sheet1')
            if not isinstance(sheet1_data, list):
                sheet1_data = []
                
            for item in sheet1_data:
                if isinstance(item, dict):
                    ws.cell(row=row, column=1, value=item.get('company_name', ''))
                    ws.cell(row=row, column=2, value=item.get('oil_depots_count', 0))
                    ws.cell(row=row, column=3, value=item.get('azs_count', 0))
                    ws.cell(row=row, column=4, value=item.get('working_azs_count', 0))
                    ws.cell(row=row, column=5, value=item.get('affiliation', ''))
                    row += 1
            
            row += 2  # Пропуск строки между компаниями
        
        # Автоподбор ширины
        self._auto_adjust_columns(ws)
        
        self._apply_default_styles(ws)
    
    def _create_structure_sheet(self, wb, data, report_date):
        """Лист структуры (аналог листа 1)"""
        ws = wb.create_sheet(title="Структура")
        
        ws['A1'] = "СТРУКТУРА ТОПЛИВНОГО РЫНКА"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        headers = ["Принадлежность", "Наименование компании", "Кол-во нефтебаз", 
                  "Кол-во АЗС", "Кол-во работающих АЗС"]
        
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row += 1
        
        # Собираем все данные из всех компаний
        all_data = []
        for company_name, company_data in data.items():
            if isinstance(company_data, dict):
                sheet1_data = company_data.get('sheet1')
                if isinstance(sheet1_data, list):
                    for item in sheet1_data:
                        if isinstance(item, dict):
                            all_data.append(item)
        
        # Сортируем по принадлежности
        all_data.sort(key=lambda x: x.get('affiliation', ''))
        
        for item in all_data:
            ws.cell(row=row, column=1, value=item.get('affiliation', ''))
            ws.cell(row=row, column=2, value=item.get('company_name', ''))
            ws.cell(row=row, column=3, value=item.get('oil_depots_count', 0))
            ws.cell(row=row, column=4, value=item.get('azs_count', 0))
            ws.cell(row=row, column=5, value=item.get('working_azs_count', 0))
            row += 1
        
        # Автоподбор ширины
        self._auto_adjust_columns(ws)
        
        self._apply_default_styles(ws)
    
    def _create_demand_sheet(self, wb, data, report_date):
        """Лист потребности (аналог листа 2)"""
        ws = wb.create_sheet(title="Потребность")
        
        ws['A1'] = "ПОТРЕБНОСТЬ В МОТОРНОМ ТОПЛИВЕ"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:K1')
        
        # Годовая потребность
        ws['A3'] = "ГОДОВАЯ ПОТРЕБНОСТЬ"
        ws['A3'].font = Font(bold=True)
        ws.merge_cells('A3:K3')
        
        headers_year = ["Компания", "Бензин всего", "АИ-76/80", "АИ-92", "АИ-95", "АИ-98/100",
                       "Дизель всего", "Зимнее", "Арктическое", "Летнее", "Межсезонное"]
        
        row = 5
        for col, header in enumerate(headers_year, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        row += 1
        
        for company_name, company_data in data.items():
            if not isinstance(company_data, dict):
                continue
                
            sheet2_data = company_data.get('sheet2', {})
            if not isinstance(sheet2_data, dict):
                sheet2_data = {}
                
            ws.cell(row=row, column=1, value=company_name)
            
            # Числовые данные с форматированием
            self._format_number_cell(ws.cell(row=row, column=2), sheet2_data.get('gasoline_total', 0))
            self._format_number_cell(ws.cell(row=row, column=3), sheet2_data.get('gasoline_ai76_80', 0))
            self._format_number_cell(ws.cell(row=row, column=4), sheet2_data.get('gasoline_ai92', 0))
            self._format_number_cell(ws.cell(row=row, column=5), sheet2_data.get('gasoline_ai95', 0))
            self._format_number_cell(ws.cell(row=row, column=6), sheet2_data.get('gasoline_ai98_100', 0))
            self._format_number_cell(ws.cell(row=row, column=7), sheet2_data.get('diesel_total', 0))
            self._format_number_cell(ws.cell(row=row, column=8), sheet2_data.get('diesel_winter', 0))
            self._format_number_cell(ws.cell(row=row, column=9), sheet2_data.get('diesel_arctic', 0))
            self._format_number_cell(ws.cell(row=row, column=10), sheet2_data.get('diesel_summer', 0))
            self._format_number_cell(ws.cell(row=row, column=11), sheet2_data.get('diesel_intermediate', 0))
            
            row += 1
        
        # Итоговая строка
        ws.cell(row=row, column=1, value="ИТОГО")
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        # Формулы для суммирования
        for col in range(2, 12):
            start_row = 6
            end_row = row - 1
            formula = f"=SUM({get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row})"
            ws.cell(row=row, column=col, value=formula)
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        self._apply_default_styles(ws)
        self._auto_adjust_columns(ws)
    
    def _create_supply_summary_sheet(self, wb, data, report_date):
        """Свод по поставкам (лист 4)"""
        ws = wb.create_sheet(title="Поставки")
        
        ws['A1'] = "СВОДНЫЕ ПОСТАВКИ ТОПЛИВА"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        headers = ["Компания", "Поставки АИ-92", "Поставки АИ-95", 
                  "Поставки дизеля зим.", "Всего поставки"]
        
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        row += 1
        
        for company_name, company_data in data.items():
            if not isinstance(company_data, dict):
                continue
                
            # Получаем данные о поставках (пока заглушка)
            ws.cell(row=row, column=1, value=company_name)
            
            # Заглушки для поставок (когда будут данные из БД)
            supply_ai92 = 0
            supply_ai95 = 0
            supply_diesel_winter = 0
            
            # Числовые данные с форматированием
            self._format_number_cell(ws.cell(row=row, column=2), supply_ai92)
            self._format_number_cell(ws.cell(row=row, column=3), supply_ai95)
            self._format_number_cell(ws.cell(row=row, column=4), supply_diesel_winter)
            
            # Формула для суммы
            total_supply = f"=SUM(B{row}:D{row})"
            ws.cell(row=row, column=5, value=total_supply)
            
            row += 1
        
        # Итоговая строка
        ws.cell(row=row, column=1, value="ИТОГО")
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        # Формулы для суммирования
        for col in range(2, 5):
            start_row = 4
            end_row = row - 1
            formula = f"=SUM({get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row})"
            ws.cell(row=row, column=col, value=formula)
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        # Формула для итоговой поставки
        ws.cell(row=row, column=5, value=f"=SUM(B{row}:D{row})")
        
        self._apply_default_styles(ws)
        self._auto_adjust_columns(ws)
    
    def _create_balance_summary_sheet(self, wb, data, report_date):
        """Свод по остаткам"""
        ws = wb.create_sheet(title="Остатки_свод")
        
        ws['A1'] = "СВОДНЫЕ ОСТАТКИ ТОПЛИВА"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        headers = ["Компания", "АИ-92", "АИ-95", "Дизель зимнее", 
                  "Дизель арктическое", "Всего дизель"]
        
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        row += 1
        
        for company_name, company_data in data.items():
            if not isinstance(company_data, dict):
                continue
                
            sheet3_totals = company_data.get('sheet3_totals', {})
            if not isinstance(sheet3_totals, dict):
                sheet3_totals = {}
                
            ws.cell(row=row, column=1, value=company_name)
            
            # Числовые данные с форматированием
            self._format_number_cell(ws.cell(row=row, column=2), sheet3_totals.get('total_stock_ai92', 0))
            self._format_number_cell(ws.cell(row=row, column=3), sheet3_totals.get('total_stock_ai95', 0))
            self._format_number_cell(ws.cell(row=row, column=4), sheet3_totals.get('total_stock_diesel_winter', 0))
            self._format_number_cell(ws.cell(row=row, column=5), sheet3_totals.get('total_stock_diesel_arctic', 0))
            
            # Формула для суммы дизеля
            diesel_total_cell = f"=D{row}+E{row}"
            ws.cell(row=row, column=6, value=diesel_total_cell)
            
            row += 1
        
        # Итоговая строка
        ws.cell(row=row, column=1, value="ИТОГО")
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        # Формулы для итогов
        for col in range(2, 6):
            start_cell = ws.cell(row=4, column=col).coordinate
            end_cell = ws.cell(row=row-1, column=col).coordinate
            ws.cell(row=row, column=col, value=f"=SUM({start_cell}:{end_cell})")
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        # Формула для итогового дизеля
        ws.cell(row=row, column=6, value=f"=D{row}+E{row}")
        
        # Форматирование итоговых числовых ячеек
        for col in range(2, 6):
            ws.cell(row=row, column=col).number_format = '0.00'
        
        self._apply_default_styles(ws)
        self._auto_adjust_columns(ws)
    
    def _create_sales_summary_sheet(self, wb, data, report_date):
        """Свод по реализации"""
        ws = wb.create_sheet(title="Реализация_свод")
        
        ws['A1'] = "СВОДНАЯ РЕАЛИЗАЦИЯ ТОПЛИВА (с начала месяца)"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        headers = ["Компания", "АИ-92", "АИ-95", "Дизель зимнее", 
                  "Дизель арктическое", "Всего реализация"]
        
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        
        row += 1
        
        for company_name, company_data in data.items():
            if not isinstance(company_data, dict):
                continue
                
            sheet5_totals = company_data.get('sheet5_totals', {})
            if not isinstance(sheet5_totals, dict):
                sheet5_totals = {}
                
            ws.cell(row=row, column=1, value=company_name)
            
            # Числовые данные с форматированием
            self._format_number_cell(ws.cell(row=row, column=2), sheet5_totals.get('total_monthly_ai92', 0))
            self._format_number_cell(ws.cell(row=row, column=3), sheet5_totals.get('total_monthly_ai95', 0))
            self._format_number_cell(ws.cell(row=row, column=4), sheet5_totals.get('total_monthly_diesel_winter', 0))
            self._format_number_cell(ws.cell(row=row, column=5), sheet5_totals.get('total_monthly_diesel_arctic', 0))
            
            # Формула для суммы реализации
            sales_total_cell = f"=SUM(B{row}:E{row})"
            ws.cell(row=row, column=6, value=sales_total_cell)
            
            row += 1
        
        # Итоговая строка
        ws.cell(row=row, column=1, value="ИТОГО")
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        # Формулы для итогов
        for col in range(2, 6):
            start_cell = ws.cell(row=4, column=col).coordinate
            end_cell = ws.cell(row=row-1, column=col).coordinate
            ws.cell(row=row, column=col, value=f"=SUM({start_cell}:{end_cell})")
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        # Формула для итоговой реализации
        ws.cell(row=row, column=6, value=f"=SUM(B{row}:E{row})")
        
        # Форматирование итоговых числовых ячеек
        for col in range(2, 6):
            ws.cell(row=row, column=col).number_format = '0.00'
        
        self._apply_default_styles(ws)
        self._auto_adjust_columns(ws)

    def generate_summary_report(self, report_date: datetime = None) -> str:
        """Генерация полного сводного отчета"""
        if report_date is None:
            report_date = datetime.now().date()
        else:
            report_date = report_date.date() if isinstance(report_date, datetime) else report_date
        
        # Получаем агрегированные данные
        aggregated_data = self.db.get_aggregated_data(report_date)
        
        if not aggregated_data:
            raise Exception("Нет данных для генерации отчета")
        
        print(f"Генерация полного отчета на дату: {report_date}")
        print(f"Найдено компаний: {len(aggregated_data)}")
        
        # Создаем книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводный отчет"
        
        # Заголовок
        ws['A1'] = "ПОЛНЫЙ СВОДНЫЙ ОТЧЕТ ПО ТОПЛИВООБЕСПЕЧЕНИЮ"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Дата отчета: {report_date.strftime('%d.%m.%Y')}"
        ws['A2'].font = Font(bold=True)
        ws.merge_cells('A2:E2')
        
        ws['A3'] = f"Количество компаний: {len(aggregated_data)}"
        ws.merge_cells('A3:E3')
        
        # Заголовки таблицы
        headers = ["№", "Компания", "АЗС", "Остатки АИ-92 (т)", 
                "Остатки АИ-95 (т)", "Реализация АИ-92 (т/мес)"]
        
        row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Данные по компаниям
        for idx, (company_name, company_data) in enumerate(aggregated_data.items(), 1):
            if not isinstance(company_data, dict):
                continue
                
            # Суммируем данные по АЗС
            azs_count = 0
            sheet1_data = company_data.get('sheet1', [])
            if isinstance(sheet1_data, list):
                for item in sheet1_data:
                    if isinstance(item, dict):
                        azs_count += item.get('azs_count', 0)
            
            # Остатки АИ-92, АИ-95
            stock_ai92 = 0
            stock_ai95 = 0
            sheet3_totals = company_data.get('sheet3_totals', {})
            if isinstance(sheet3_totals, dict):
                stock_ai92 = sheet3_totals.get('total_stock_ai92', 0)
                stock_ai95 = sheet3_totals.get('total_stock_ai95', 0)
            
            # Реализация АИ-92
            sales_ai92 = 0
            sheet5_totals = company_data.get('sheet5_totals', {})
            if isinstance(sheet5_totals, dict):
                sales_ai92 = sheet5_totals.get('total_monthly_ai92', 0)
            
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=company_name)
            ws.cell(row=row, column=3, value=azs_count)
            ws.cell(row=row, column=4, value=stock_ai92)
            ws.cell(row=row, column=5, value=stock_ai95)
            ws.cell(row=row, column=6, value=sales_ai92)
            
            # Форматирование числовых ячеек
            for col in [4, 5, 6]:
                ws.cell(row=row, column=col).number_format = '0.00'
            
            row += 1
        
        # Итоговая строка
        ws.cell(row=row, column=1, value="ИТОГО:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        # Формулы для итогов
        for col in range(3, 7):
            start_row = 6
            end_row = row - 1
            ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row})")
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        # Автоподбор ширины
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Добавляем границы
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        # Сохраняем файл
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        reports_folder = 'reports_output'
        os.makedirs(reports_folder, exist_ok=True)
        output_path = os.path.join(reports_folder, f'Полный_сводный_отчет_{timestamp}.xlsx')
        
        wb.save(output_path)
        
        print(f"Отчет сохранен: {output_path}")
        return output_path

    def generate_summary_report_with_data(self, data: Dict[str, Any], report_date: datetime = None) -> str:
        """Генерация отчета с готовыми данными"""
        if report_date is None:
            report_date = datetime.now().date()
        else:
            report_date = report_date.date() if isinstance(report_date, datetime) else report_date
        
        print(f"Генерация отчета с {len(data)} компаниями на дату: {report_date}")
        
        if not data:
            raise Exception("Нет данных для генерации отчета")
        
        # Создаем книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводный отчет"
        
        # Заголовок
        ws['A1'] = "ПОЛНЫЙ СВОДНЫЙ ОТЧЕТ ПО ТОПЛИВООБЕСПЕЧЕНИЮ"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Дата отчета: {report_date.strftime('%d.%m.%Y')}"
        ws['A2'].font = Font(bold=True)
        ws.merge_cells('A2:F2')
        
        ws['A3'] = f"Количество компаний: {len(data)}"
        ws.merge_cells('A3:F3')
        
        # Заголовки таблицы
        headers = ["№", "Компания", "АЗС", "Остатки АИ-92 (т)", 
                "Остатки АИ-95 (т)", "Реализация АИ-92 (т/мес)"]
        
        row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Данные по компаниям
        for idx, (company_name, company_data) in enumerate(data.items(), 1):
            # Суммируем данные по АЗС
            azs_count = 0
            sheet1_data = company_data.get('sheet1', [])
            if isinstance(sheet1_data, list):
                for item in sheet1_data:
                    if isinstance(item, dict):
                        azs_count += item.get('azs_count', 0)
            
            # Остатки АИ-92, АИ-95
            stock_ai92 = company_data.get('sheet3_totals', {}).get('total_stock_ai92', 0)
            stock_ai95 = company_data.get('sheet3_totals', {}).get('total_stock_ai95', 0)
            
            # Реализация АИ-92
            sales_ai92 = company_data.get('sheet5_totals', {}).get('total_monthly_ai92', 0)
            
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=company_name)
            ws.cell(row=row, column=3, value=azs_count)
            ws.cell(row=row, column=4, value=stock_ai92)
            ws.cell(row=row, column=5, value=stock_ai95)
            ws.cell(row=row, column=6, value=sales_ai92)
            
            # Форматирование числовых ячеек
            for col in [4, 5, 6]:
                ws.cell(row=row, column=col).number_format = '0.00'
            
            row += 1
        
        # Итоговая строка
        ws.cell(row=row, column=1, value="ИТОГО:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        # Формулы для итогов
        for col in range(3, 7):
            start_row = 6
            end_row = row - 1
            ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row})")
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        # Автоподбор ширины
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Добавляем границы
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        for row_cells in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row_cells:
                cell.border = thin_border
        
        # Сохраняем файл
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        reports_folder = 'reports_output'
        os.makedirs(reports_folder, exist_ok=True)
        output_path = os.path.join(reports_folder, f'Полный_сводный_отчет_{timestamp}.xlsx')
        
        wb.save(output_path)
        
        print(f"Отчет сохранен: {output_path}")
        return output_path

    def generate_simple_report(self, report_date: datetime = None) -> str:
        """Генерация простого отчета для тестирования"""
        if report_date is None:
            report_date = datetime.now().date()
        else:
            report_date = report_date.date() if isinstance(report_date, datetime) else report_date
        
        # Получаем агрегированные данные
        aggregated_data = self.db.get_aggregated_data(report_date)
        
        if not aggregated_data:
            raise Exception("Нет данных для генерации отчета")
        
        # Создаем простую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводный отчет"
        
        # Заголовок
        ws['A1'] = "СВОДНЫЙ ОТЧЕТ ПО ТОПЛИВООБЕСПЕЧЕНИЮ"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        
        ws['A2'] = f"Дата отчета: {report_date.strftime('%d.%m.%Y')}"
        ws['A2'].font = Font(bold=True)
        ws.merge_cells('A2:C2')
        
        # Заголовки таблицы
        headers = ["Компания", "Кол-во АЗС", "Остатки АИ-92 (т)"]
        row = 4
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
        
        row += 1
        
        # Данные
        for company_name, company_data in aggregated_data.items():
            if not isinstance(company_data, dict):
                continue
                
            # Суммируем данные по АЗС
            azs_count = 0
            sheet1_data = company_data.get('sheet1', [])
            if isinstance(sheet1_data, list):
                for item in sheet1_data:
                    if isinstance(item, dict):
                        azs_count += item.get('azs_count', 0)
            
            # Остатки АИ-92
            stock_ai92 = 0
            sheet3_totals = company_data.get('sheet3_totals', {})
            if isinstance(sheet3_totals, dict):
                stock_ai92 = sheet3_totals.get('total_stock_ai92', 0)
            
            ws.cell(row=row, column=1, value=company_name)
            ws.cell(row=row, column=2, value=azs_count)
            ws.cell(row=row, column=3, value=stock_ai92)
            
            row += 1
        
        # Автоподбор ширины
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        reports_folder = 'reports_output'
        os.makedirs(reports_folder, exist_ok=True)
        output_path = os.path.join(reports_folder, f'Сводный_отчет_{timestamp}.xlsx')
        
        wb.save(output_path)
        
        return output_path
    
class TemplateReportGenerator:
    """Генератор отчетов на основе шаблона Excel"""
    
    def __init__(self, db_connection, template_path: str = None):
        self.db = db_connection
        self.template_filler = TemplateFiller(template_path)
    
    def generate_from_template(self, report_date: datetime = None, 
                             output_path: str = None) -> str:
        """Генерация отчета из шаблона"""
        if report_date is None:
            report_date = datetime.now().date()
        
        print(f"Генерация отчета из шаблона на дату: {report_date}")
        
        # Получаем агрегированные данные
        aggregated_data = self.db.get_aggregated_data(report_date)
        
        if not aggregated_data:
            # Пробуем найти данные за последние дни
            for days_back in range(1, 4):
                test_date = report_date - timedelta(days=days_back)
                aggregated_data = self.db.get_aggregated_data(test_date)
                if aggregated_data:
                    report_date = test_date
                    print(f"Используем данные на дату: {report_date}")
                    break
        
        if not aggregated_data:
            raise Exception("Нет данных для генерации отчета")
        
        # Подготавливаем данные для шаблона
        template_data = self._prepare_template_data(aggregated_data, report_date)
        
        # Заполняем шаблон
        report_path = self.template_filler.fill_template(template_data, output_path)
        
        return report_path

    def _prepare_template_data(self, aggregated_data: Dict[str, Any], 
                            report_date: datetime) -> Dict[str, Any]:
        """Подготовка данных для шаблона"""
        companies = []
        
        for company_name, company_data in aggregated_data.items():
            print(f"Подготовка данных для компании: {company_name}")
            
            # Подсчитываем общее количество АЗС
            azs_count = 0
            working_azs = 0
            for item in company_data.get('sheet1', []):
                azs_count += item.get('azs_count', 0)
                working_azs += item.get('working_azs_count', 0)
            
            # Получаем потребность
            demand_data = company_data.get('sheet2', {})
            
            # Получаем остатки
            stock_totals = company_data.get('sheet3_totals', {})
            
            # Получаем поставки (данные из sheet4)
            supply_data = company_data.get('sheet4_data', [])
            total_supply_ai92 = sum(item.get('supply_ai92', 0) for item in supply_data)
            total_supply_ai95 = sum(item.get('supply_ai95', 0) for item in supply_data)
            total_supply_diesel_winter = sum(item.get('supply_diesel_winter', 0) for item in supply_data)
            total_supply_diesel_arctic = sum(item.get('supply_diesel_arctic', 0) for item in supply_data)
            
            # Получаем реализацию
            sales_totals = company_data.get('sheet5_totals', {})
            
            company_info = {
                'name': company_name,
                'azs_count': azs_count or 0,
                'working_azs': working_azs or 0,
                'gasoline_demand': demand_data.get('gasoline_total', 0) or 0,
                'diesel_demand': demand_data.get('diesel_total', 0) or 0,
                'monthly_gasoline': demand_data.get('monthly_gasoline_total', 0) or 0,
                'monthly_diesel': demand_data.get('monthly_diesel_total', 0) or 0,
                'stock_ai92': stock_totals.get('total_stock_ai92', 0) or 0,
                'stock_ai95': stock_totals.get('total_stock_ai95', 0) or 0,
                'stock_diesel_winter': stock_totals.get('total_stock_diesel_winter', 0) or 0,
                'stock_diesel_arctic': stock_totals.get('total_stock_diesel_arctic', 0) or 0,
                'supply_ai92': total_supply_ai92 or 0,
                'supply_ai95': total_supply_ai95 or 0,
                'supply_diesel_winter': total_supply_diesel_winter or 0,
                'supply_diesel_arctic': total_supply_diesel_arctic or 0,
                'sales_ai92': sales_totals.get('total_monthly_ai92', 0) or 0,
                'sales_ai95': sales_totals.get('total_monthly_ai95', 0) or 0,
                'sales_diesel_winter': sales_totals.get('total_monthly_diesel_winter', 0) or 0,
                'sales_diesel_arctic': sales_totals.get('total_monthly_diesel_arctic', 0) or 0,
            }
            
            print(f"  АЗС: {azs_count}, Поставки AI92: {company_info['supply_ai92']}, Продажи AI92: {company_info['sales_ai92']}")
            
            companies.append(company_info)
        
        template_data = {
            'report_date': report_date,
            'companies': companies,
            'total_companies': len(companies)
        }
        
        print(f"Всего подготовлено компаний: {len(companies)}")
        return template_data
    
# В report_generator.py добавьте новый класс:

class SimpleTemplateReportGenerator:
    """Простой генератор отчетов на основе шаблона Excel (только цифры)"""
    
    def __init__(self, db_connection, template_path: str = None):
        self.db = db_connection
        
        try:
            from reports.simple_template_filler import SimpleTemplateFiller
            self.template_filler = SimpleTemplateFiller(template_path)
        except ImportError as e:
            print(f"ВНИМАНИЕ: SimpleTemplateFiller не найден: {e}")
            self.template_filler = None
    
    def generate_from_template(self, report_date: datetime = None, 
                             output_path: str = None) -> str:
        """Генерация отчета из шаблона (простая версия)"""
        if self.template_filler is None:
            raise Exception("SimpleTemplateFiller не доступен")
        
        if report_date is None:
            report_date = datetime.now().date()
        
        print(f"Генерация отчета из шаблона на дату: {report_date}")
        
        # Получаем агрегированные данные
        aggregated_data = self.db.get_aggregated_data(report_date)
        
        if not aggregated_data:
            # Пробуем найти данные за последние дни
            for days_back in range(1, 4):
                test_date = report_date - timedelta(days=days_back)
                aggregated_data = self.db.get_aggregated_data(test_date)
                if aggregated_data:
                    report_date = test_date
                    print(f"Используем данные на дату: {report_date}")
                    break
        
        if not aggregated_data:
            # Берем все данные без фильтра по дате
            aggregated_data = self.db.get_aggregated_data()
            print("Используем все данные без фильтра по дате")
        
        if not aggregated_data:
            raise Exception("Нет данных для генерации отчета")
        
        print(f"Найдено компаний: {len(aggregated_data)}")
        
        # Подготавливаем данные для шаблона
        template_data = self._prepare_simple_data(aggregated_data, report_date)
        
        # Заполняем шаблон
        report_path = self.template_filler.fill_template(template_data, output_path)
        
        return report_path
    
    def _prepare_simple_data(self, aggregated_data: Dict[str, Any], 
                           report_date: datetime) -> Dict[str, Any]:
        """Подготовка простых данных для шаблона"""
        companies = []
        
        for company_name, company_data in aggregated_data.items():
            print(f"Подготовка данных для компании: {company_name}")
            
            # Берем только основные цифровые данные
            demand_data = company_data.get('sheet2', {})
            stock_totals = company_data.get('sheet3_totals', {})
            sales_totals = company_data.get('sheet5_totals', {})
            
            # Данные поставок (из sheet4)
            supply_data = company_data.get('sheet4_data', [])
            total_supply_ai92 = sum(item.get('supply_ai92', 0) for item in supply_data)
            total_supply_ai95 = sum(item.get('supply_ai95', 0) for item in supply_data)
            total_supply_diesel_winter = sum(item.get('supply_diesel_winter', 0) for item in supply_data)
            total_supply_diesel_arctic = sum(item.get('supply_diesel_arctic', 0) for item in supply_data)
            
            company_info = {
                'name': company_name,
                'gasoline_demand': demand_data.get('gasoline_total', 0) or 0,
                'diesel_demand': demand_data.get('diesel_total', 0) or 0,
                'monthly_gasoline': demand_data.get('monthly_gasoline_total', 0) or 0,
                'monthly_diesel': demand_data.get('monthly_diesel_total', 0) or 0,
                'stock_ai92': stock_totals.get('total_stock_ai92', 0) or 0,
                'stock_ai95': stock_totals.get('total_stock_ai95', 0) or 0,
                'stock_diesel_winter': stock_totals.get('total_stock_diesel_winter', 0) or 0,
                'stock_diesel_arctic': stock_totals.get('total_stock_diesel_arctic', 0) or 0,
                'supply_ai92': total_supply_ai92 or 0,
                'supply_ai95': total_supply_ai95 or 0,
                'supply_diesel_winter': total_supply_diesel_winter or 0,
                'supply_diesel_arctic': total_supply_diesel_arctic or 0,
                'sales_ai92': sales_totals.get('total_monthly_ai92', 0) or 0,
                'sales_ai95': sales_totals.get('total_monthly_ai95', 0) or 0,
                'sales_diesel_winter': sales_totals.get('total_monthly_diesel_winter', 0) or 0,
                'sales_diesel_arctic': sales_totals.get('total_monthly_diesel_arctic', 0) or 0,
            }
            
            print(f"  Данные: AI92 остатки={company_info['stock_ai92']}, поставки={company_info['supply_ai92']}, продажи={company_info['sales_ai92']}")
            
            companies.append(company_info)
        
        template_data = {
            'report_date': report_date,
            'companies': companies,
            'total_companies': len(companies)
        }
        
        print(f"Всего подготовлено компаний: {len(companies)}")
        return template_data