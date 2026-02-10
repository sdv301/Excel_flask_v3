# reports/template_report_generator.py - ПОЛНАЯ ВЕРСИЯ БЕЗ ОГРАНИЧЕНИЙ
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import json

class TemplateReportGenerator:
    def __init__(self, db_connection, template_path: str = None):
        self.db = db_connection
        
        # Всегда сохраняем в reports_output относительно корня проекта
        self.reports_dir = 'reports_output'
        
        # Создаем папку если не существует
        os.makedirs(self.reports_dir, exist_ok=True)
        
        # Путь к шаблону
        if template_path is None:
            self.template_path = 'report_templates/Сводный_отчет_шаблон.xlsx'
        else:
            self.template_path = template_path
        
        if not os.path.exists(self.template_path):
            # Попробуем найти шаблон в других местах
            possible_paths = [
                'report_templates/Сводный_отчет_шаблон.xlsx',
                '../report_templates/Сводный_отчет_шаблон.xlsx',
                './report_templates/Сводный_отчет_шаблон.xlsx'
            ]
            
            for path in possible_paths:
                if os.path.exists(path):
                    self.template_path = path
                    break
            else:
                raise FileNotFoundError(f"Шаблон не найден. Искал: {possible_paths}")

    def generate_report(self, report_date: date = None) -> str:
        try:
            if report_date is None:
                report_date = datetime.now().date()

            print(f"\n🎯 ГЕНЕРАЦИЯ ОТЧЕТА НА {report_date.strftime('%d.%m.%Y')}")

            # Получаем данные из БД
            aggregated_data = self.db.get_aggregated_data()
            if not aggregated_data:
                raise Exception("Нет данных в БД")

            # Создаем имя файла
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'Сводный_отчет_{timestamp}.xlsx'
            output_path = os.path.join(self.reports_dir, filename)
            
            print(f"📁 Сохраняем в: {output_path}")
            print(f"📁 Абсолютный путь: {os.path.abspath(output_path)}")

            # Копируем шаблон
            shutil.copy2(self.template_path, output_path)

            # Загружаем и заполняем Excel
            wb = load_workbook(output_path)
            self._update_report_info(wb, report_date, aggregated_data)
            self._fill_all_company_data(wb, aggregated_data)
            wb.save(output_path)

            # Проверяем что файл создан
            if os.path.exists(output_path):
                size = os.path.getsize(output_path)
                print(f"✅ Отчет создан успешно! Размер: {size} байт")
                return output_path
            else:
                raise Exception("Файл не был создан")
                
        except Exception as e:
            print(f"❌ Ошибка: {e}")
            raise

    def _update_report_info(self, wb, report_date: date, aggregated_data: dict):
        """Обновление общей информации в отчете"""
        date_str = report_date.strftime('%d.%m.%Y')
        
        # Обновляем дату в шапке отчета
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Ищем ячейку с датой (обычно в первых 5 строках)
            for row in range(1, 6):
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and 'дата' in str(cell.value).lower():
                        # Записываем дату в соседнюю ячейку
                        ws.cell(row=row, column=col+1).value = date_str
                        print(f"📅 Обновлена дата в {sheet_name}: {date_str}")

    def _fill_all_company_data(self, wb, aggregated_data: dict):
        """Заполнение ВСЕХ данных по компаниям во все разделы"""
        
        print(f"\n🏢 ОБРАБОТКА ВСЕХ ЛИСТОВ ОТЧЕТА:")
        
        # Лист 1: Структура компаний - ВСЕ данные
        # if '1-Структура' in wb.sheetnames:
        #     self._fill_structure_sheet_full(wb['1-Структура'], aggregated_data)
        
        # # Лист 2: Потребность - ВСЕ данные
        # if '2-Потребность' in wb.sheetnames:
        #     self._fill_demand_sheet_full(wb['2-Потребность'], aggregated_data)
        
        # Лист 3: Остатки - ВСЕ данные
        if '3-Остатки' in wb.sheetnames:
            self._fill_stocks_sheet_full(wb['3-Остатки'], aggregated_data)
        
        # Лист 4: Поставки - ВСЕ данные  
        if '4-Поставка' in wb.sheetnames:
            self._fill_supply_sheet_full(wb['4-Поставка'], aggregated_data)
        
        # Лист 5: Реализация - ВСЕ данные
        if '5-Реализация' in wb.sheetnames:
            self._fill_sales_sheet_full(wb['5-Реализация'], aggregated_data)
        
        # Лист 6: Авиатопливо - если есть данные
        if '6-Авиатопливо' in wb.sheetnames:
            self._fill_aviation_sheet(wb['6-Авиатопливо'], aggregated_data)
        
        # # Лист 7: Комментарии/Справка - стандартные комментарии
        # sheet7_name = self._get_sheet7_name(wb)
        # if sheet7_name:
        #     self._fill_comments_sheet(wb[sheet7_name], aggregated_data)

    # def _fill_structure_sheet_full(self, ws, aggregated_data: dict):
    #     """Заполнение листа со структурой компаний - ВСЕ данные"""
    #     print(f"📋 Заполнение листа 'Структура' (все данные)...")
        
    #     # Определяем стартовую строку для данных
    #     start_row = 13
    #     current_row = start_row
        
    #     for company_name, company_data in aggregated_data.items():
    #         sheet1_data = company_data.get('sheet1', [])
            
    #         if sheet1_data:
    #             # Выводим ВСЕ записи для компании (кроме заголовков)
    #             for record in sheet1_data:
    #                 # Пропускаем строки-заголовки
    #                 if (record.get('company_name') and 
    #                     'наименование компаний' in str(record.get('company_name')).lower()):
    #                     continue
                    
    #                 if record.get('company_name') == '2':  # Пропускаем технические строки
    #                     continue
                        
    #                 # Заполняем данные в строку
    #                 self._set_cell_value(ws, current_row, 1, record.get('affiliation', ''))
    #                 self._set_cell_value(ws, current_row, 2, record.get('company_name', company_name))
    #                 self._set_cell_value(ws, current_row, 3, record.get('oil_depots_count', 0))
    #                 self._set_cell_value(ws, current_row, 4, record.get('azs_count', 0))
    #                 self._set_cell_value(ws, current_row, 5, record.get('working_azs_count', 0))
                    
    #                 current_row += 1
        
    #     print(f"  ✅ Выгружено записей структуры: {current_row - start_row}")

    # def _fill_demand_sheet_full(self, ws, aggregated_data: dict):
    #     """Заполнение листа с потребностью - ВСЕ данные"""
    #     print(f"📈 Заполнение листа 'Потребность' (все данные)...")
        
    #     # Определяем строки для данных (из вашего шаблона)
    #     year_row = 7
    #     month_row = 13
        
    #     total_companies = len(aggregated_data)
        
    #     # Если компаний несколько, распределяем данные по строкам
    #     current_year_row = year_row
    #     current_month_row = month_row
        
    #     for company_name, company_data in aggregated_data.items():
    #         sheet2_data = company_data.get('sheet2', {})
            
    #         if sheet2_data:
    #             # Годовые данные
    #             self._set_cell_value(ws, current_year_row, 1, company_name)  # Название компании
    #             self._set_cell_value(ws, current_year_row, 4, round(sheet2_data.get('gasoline_ai92', 0), 3))
    #             self._set_cell_value(ws, current_year_row, 5, round(sheet2_data.get('gasoline_ai95', 0), 3))
    #             self._set_cell_value(ws, current_year_row, 8, round(sheet2_data.get('diesel_winter', 0), 3))
    #             self._set_cell_value(ws, current_year_row, 9, round(sheet2_data.get('diesel_arctic', 0), 3))
                
    #             # Месячные данные
    #             self._set_cell_value(ws, current_month_row, 1, company_name)  # Название компании
    #             self._set_cell_value(ws, current_month_row, 4, round(sheet2_data.get('monthly_gasoline_total', 0) / 2, 3))
    #             self._set_cell_value(ws, current_month_row, 5, round(sheet2_data.get('monthly_gasoline_total', 0) / 2, 3))
    #             self._set_cell_value(ws, current_month_row, 8, round(sheet2_data.get('monthly_diesel_total', 0) / 2, 3))
    #             self._set_cell_value(ws, current_month_row, 9, round(sheet2_data.get('monthly_diesel_total', 0) / 2, 3))
                
    #             current_year_row += 1
    #             current_month_row += 1
        
    #     print(f"  ✅ Заполнены данные потребности для {total_companies} компаний")

    def _fill_stocks_sheet_full(self, ws, aggregated_data: dict):
        """Заполнение листа с остатками - ВСЕ данные"""
        print(f"📦 Заполнение листа 'Остатки' (все данные)...")
        
        start_row = 9
        current_row = start_row
        total_locations = 0
        
        for company_name, company_data in aggregated_data.items():
            sheet3_data = company_data.get('sheet3_data', [])
            
            # Записываем ВСЕ локации компании
            for location_data in sheet3_data:
                self._set_cell_value(ws, current_row, 2, company_name)  # Компания
                self._set_cell_value(ws, current_row, 3, location_data.get('location_name', ''))
                
                self._set_cell_value(ws, current_row, 5, round(location_data.get('stock_ai92', 0), 3))
                self._set_cell_value(ws, current_row, 6, round(location_data.get('stock_ai95', 0), 3))
                self._set_cell_value(ws, current_row, 7, round(location_data.get('stock_ai98_ai100', 0), 3))
                self._set_cell_value(ws, current_row, 8, round(location_data.get('stock_diesel_winter', 0), 3))
                self._set_cell_value(ws, current_row, 9, round(location_data.get('stock_diesel_arctic', 0), 3))
                self._set_cell_value(ws, current_row, 10, round(location_data.get('stock_diesel_summer', 0), 3))
                
                self._set_cell_value(ws, current_row, 13, round(location_data.get('transit_ai92', 0), 3))
                self._set_cell_value(ws, current_row, 14, round(location_data.get('transit_ai95', 0), 3))
                self._set_cell_value(ws, current_row, 15, round(location_data.get('transit_ai98_ai100', 0), 3))
                self._set_cell_value(ws, current_row, 16, round(location_data.get('transit_diesel_winter', 0), 3))
                self._set_cell_value(ws, current_row, 17, round(location_data.get('transit_diesel_arctic', 0), 3))
                self._set_cell_value(ws, current_row, 18, round(location_data.get('transit_diesel_summer', 0), 3))
                
                self._set_cell_value(ws, current_row, 21, round(location_data.get('capacity_ai92', 0), 3))
                self._set_cell_value(ws, current_row, 22, round(location_data.get('capacity_ai95', 0), 3))
                self._set_cell_value(ws, current_row, 23, round(location_data.get('capacity_ai98_ai100', 0), 3))
                self._set_cell_value(ws, current_row, 24, round(location_data.get('capacity_diesel_winter', 0), 3))
                self._set_cell_value(ws, current_row, 25, round(location_data.get('capacity_diesel_arctic', 0), 3))
                self._set_cell_value(ws, current_row, 26, round(location_data.get('capacity_diesel_summer', 0), 3))
                
                current_row += 1
                total_locations += 1
        
        print(f"  ✅ Выгружено локаций с остатками: {total_locations}")

    def _fill_supply_sheet_full(self, ws, aggregated_data: dict):
        """Заполнение листа с поставками - ВСЕ данные"""
        print(f"🚚 Заполнение листа 'Поставка' (все данные)...")
        
        start_row = 9
        current_row = start_row
        total_supplies = 0
        
        for company_name, company_data in aggregated_data.items():
            sheet4_data = company_data.get('sheet4_data', [])
            
            # Записываем ВСЕ поставки компании
            for supply_data in sheet4_data:
                self._set_cell_value(ws, current_row, 2, company_name)
                self._set_cell_value(ws, current_row, 3, supply_data.get('oil_depot_name', ''))
                
                # Парсим дату поставки
                supply_date = supply_data.get('supply_date')
                if supply_date:
                    if isinstance(supply_date, str):
                        self._set_cell_value(ws, current_row, 4, supply_date)
                    else:
                        self._set_cell_value(ws, current_row, 4, str(supply_date))
                
                self._set_cell_value(ws, current_row, 6, round(supply_data.get('supply_ai92', 0), 3))
                self._set_cell_value(ws, current_row, 7, round(supply_data.get('supply_ai95', 0), 3))
                self._set_cell_value(ws, current_row, 8, round(supply_data.get('supply_ai98_100', 0), 3))
                self._set_cell_value(ws, current_row, 9, round(supply_data.get('supply_diesel_winter', 0), 3))
                self._set_cell_value(ws, current_row, 10, round(supply_data.get('supply_diesel_arctic', 0), 3))
                self._set_cell_value(ws, current_row, 11, round(supply_data.get('supply_diesel_summer', 0), 3))
                
                current_row += 1
                total_supplies += 1
        
        print(f"  ✅ Выгружено записей о поставках: {total_supplies}")

    def _fill_sales_sheet_full(self, ws, aggregated_data: dict):
        """Заполнение листа с реализацией - ВСЕ данные"""
        print(f"💰 Заполнение листа 'Реализация' (все данные)...")
        
        start_row = 9
        current_row = start_row
        total_sales = 0
        
        for company_name, company_data in aggregated_data.items():
            sheet5_data = company_data.get('sheet5_data', [])
            
            # Записываем ВСЕ данные по реализации
            for sales_data in sheet5_data:
                self._set_cell_value(ws, current_row, 2, company_name)
                self._set_cell_value(ws, current_row, 3, sales_data.get('location_name', ''))
                
                self._set_cell_value(ws, current_row, 5, round(sales_data.get('daily_ai92', 0), 3))
                self._set_cell_value(ws, current_row, 6, round(sales_data.get('daily_ai95', 0), 3))
                self._set_cell_value(ws, current_row, 6, round(sales_data.get('daily_ai98_100', 0), 3))
                self._set_cell_value(ws, current_row, 8, round(sales_data.get('daily_winter', 0), 3))
                self._set_cell_value(ws, current_row, 9, round(sales_data.get('daily_arctic', 0), 3))
                self._set_cell_value(ws, current_row, 10, round(sales_data.get('daily_summer', 0), 3))
                
                self._set_cell_value(ws, current_row, 13, round(sales_data.get('monthly_ai92', 0), 3))
                self._set_cell_value(ws, current_row, 14, round(sales_data.get('monthly_ai95', 0), 3))
                self._set_cell_value(ws, current_row, 15, round(sales_data.get('monthly_ai98_100', 0), 3))
                self._set_cell_value(ws, current_row, 16, round(sales_data.get('monthly_diesel_winter', 0), 3))
                self._set_cell_value(ws, current_row, 17, round(sales_data.get('monthly_diesel_arctic', 0), 3))
                self._set_cell_value(ws, current_row, 18, round(sales_data.get('monthly_diesel_summer', 0), 3))
                
                current_row += 1
                total_sales += 1
        
        print(f"  ✅ Выгружено записей реализации: {total_sales}")

    def _fill_aviation_sheet(self, ws, aggregated_data: dict):
        """Заполнение листа с авиатопливом"""
        print(f"✈️ Заполнение листа 'Авиатопливо'...")
        
        # В ваших данных пока нет информации по авиатопливу
        # Добавляем заглушку или базовую информацию
        start_row = 8
        current_row = start_row
        
        # Если в будущем появятся данные, можно будет их добавить
        has_aviation_data = False
        
        for company_name, company_data in aggregated_data.items():
            # Проверяем, есть ли данные по авиатопливу
            if 'sheet6' in company_data or 'aviation' in company_data:
                has_aviation_data = True
                # Здесь будет логика заполнения когда появятся данные
                break
        
        if not has_aviation_data:
            # Стандартная информация
            self._set_cell_value(ws, current_row, 1, "Данные по авиатопливу отсутствуют")
            self._set_cell_value(ws, current_row, 2, "В текущей версии отчетности")
            print(f"  ⚠️ Данные по авиатопливу отсутствуют - добавлена заглушка")

    # def _fill_comments_sheet(self, ws, aggregated_data: dict):
    #     """Заполнение листа с комментариями/справкой"""
    #     print(f"📝 Заполнение листа 'Комментарии'...")
        
    #     # Стандартные комментарии по ситуации
    #     comments_data = [
    #         {"fuel": "Бензин автомобильный", "situation": "Стабильная", "comment": "Обеспеченность в норме"},
    #         {"fuel": "Дизельное топливо", "situation": "Стабильная", "comment": "Обеспеченность в норме"},
    #         {"fuel": "Авиатопливо", "situation": "Стабильная", "comment": "Обеспеченность в норме"}
    #     ]
        
    #     start_row = 6
    #     current_row = start_row
        
    #     for comment in comments_data:
    #         self._set_cell_value(ws, current_row, 1, comment["fuel"])
    #         self._set_cell_value(ws, current_row, 2, comment["situation"])
    #         self._set_cell_value(ws, current_row, 3, comment["comment"])
    #         current_row += 1
        
    #     print(f"  ✅ Добавлены стандартные комментарии")

    def _set_cell_value(self, ws, row: int, col: int, value):
        """Безопасная установка значения ячейки с проверкой границ"""
        try:
            # Если значение None, преобразуем в пустую строку или 0
            if value is None:
                if isinstance(value, (int, float)):
                    value = 0
                else:
                    value = ""
            
            # Проверяем, что строка и колонка в пределах допустимого
            if row > 0 and col > 0:
                ws.cell(row=row, column=col).value = value
                return True
            else:
                print(f"  ⚠️ Некорректные координаты: строка {row}, колонка {col}")
                return False
                
        except Exception as e:
            print(f"  ❗ Ошибка записи в ячейку {get_column_letter(col)}{row}: {e}")
            return False

    def debug_template_structure(self):
        """Метод для отладки структуры шаблона"""
        wb = load_workbook(self.template_path)
        
        print(f"\n🔍 ДЕТАЛЬНАЯ СТРУКТУРА ШАБЛОНА '{self.template_path}':")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- Лист: {sheet_name} ---")
            print(f"Размер: {ws.max_row} строк, {ws.max_column} колонок")
            
            # Показываем структуру первых 15 строк
            for row in range(1, min(16, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(ws.max_column + 1, 12)):  # Первые 12 колонок
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        cell_text = str(cell.value)
                        if len(cell_text) > 20:
                            cell_text = cell_text[:17] + "..."
                        row_data.append(f"{get_column_letter(col)}: {cell_text}")
                
                if row_data:
                    print(f"Строка {row:2d}: {', '.join(row_data)}")
        
        wb.close()

# Дополнительная функция для массовой выгрузки
def generate_complete_report(db_connection, template_path=None):
    """Функция для генерации полного отчета со всеми данными"""
    generator = TemplateReportGenerator(db_connection, template_path)
    
    # Сначала покажем структуру шаблона для отладки
    generator.debug_template_structure()
    
    # Генерируем полный отчет
    return generator.generate_report()

if __name__ == "__main__":
    # Пример использования
    from database.queries import DatabaseQueries
    
    db = DatabaseQueries()
    report_path = generate_complete_report(db)
    print(f"\n🎉 Отчет готов: {report_path}")
