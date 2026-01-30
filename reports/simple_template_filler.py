# reports/simple_template_filler.py - ПРОСТОЙ ЗАПОЛНИТЕЛЬ ТОЛЬКО ЦИФР
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Any, Optional
import traceback

class SimpleTemplateFiller:
    def __init__(self, template_path: str = None):
        if template_path is None:
            self.template_path = os.path.join('report_templates', 'Сводный_отчет_шаблон.xlsx')
        else:
            self.template_path = template_path
        
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Шаблон не найден: {self.template_path}")
    
    def fill_template(self, data: Dict[str, Any], output_path: str = None) -> str:
        """Заполнение шаблона только цифровыми данными"""
        try:
            print(f"Загрузка шаблона: {self.template_path}")
            
            # Создаем копию шаблона
            import shutil
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            if output_path is None:
                output_dir = 'reports_output'
                os.makedirs(output_dir, exist_ok=True)
                output_path = os.path.join(output_dir, f'Сводный_отчет_{timestamp}.xlsx')
            
            # Копируем шаблон
            shutil.copy2(self.template_path, output_path)
            
            # Открываем для заполнения
            wb = load_workbook(output_path)
            
            print(f"Листы в шаблоне: {wb.sheetnames}")
            
            # Заполняем данные для каждой компании
            companies = data.get('companies', [])
            print(f"Компаний для заполнения: {len(companies)}")
            
            for i, company in enumerate(companies, 1):
                company_name = company.get('name', f'Компания_{i}')
                print(f"\n--- Заполнение данных для компании {i}: {company_name} ---")
                
                # Заполняем потребность (лист 2)
                if '2-Потребность' in wb.sheetnames:
                    self._fill_company_demand(wb['2-Потребность'], company, i)
                
                # Заполняем остатки (лист 3)
                if '3-Остатки' in wb.sheetnames:
                    self._fill_company_balance(wb['3-Остатки'], company, i)
                
                # Заполняем поставки (лист 4)
                if '4-Поставка' in wb.sheetnames:
                    self._fill_company_supply(wb['4-Поставка'], company, i)
                
                # Заполняем реализацию (лист 5)
                if '5-Реализация' in wb.sheetnames:
                    self._fill_company_sales(wb['5-Реализация'], company, i)
            
            # Обновляем дату отчета
            self._update_report_date(wb, data.get('report_date', datetime.now()))
            
            # Сохраняем
            wb.save(output_path)
            print(f"\n✅ Отчет заполнен и сохранен: {output_path}")
            
            return output_path
            
        except Exception as e:
            print(f"❌ Ошибка заполнения шаблона: {e}")
            traceback.print_exc()
            raise
    
    def _update_report_date(self, wb, report_date):
        """Обновление даты отчета на всех листах"""
        date_str = report_date.strftime('%d.%m.%Y') if isinstance(report_date, datetime) else str(report_date)
        
        # Ищем ячейки с датой
        date_patterns = ['дата', 'Дата', 'DATE', 'Отчет на', 'по состоянию на']
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in range(1, 50):
                for col in range(1, 20):
                    try:
                        cell = ws.cell(row=row, column=col)
                        if cell.value and isinstance(cell.value, str):
                            cell_str = str(cell.value)
                            # Если ячейка содержит указатель на дату
                            if any(pattern in cell_str for pattern in date_patterns):
                                # Пробуем найти соседнюю ячейку для даты
                                if col < 20:
                                    date_cell = ws.cell(row=row, column=col+1)
                                    if date_cell.value is None or isinstance(date_cell.value, (str, datetime)):
                                        date_cell.value = date_str
                                        print(f"  Обновлена дата в {sheet_name}: {date_cell.coordinate}")
                    except:
                        continue
    
    def _fill_company_demand(self, ws, company: Dict[str, Any], company_index: int):
        """Заполнение данных потребности для одной компании"""
        company_name = company.get('name', f'Компания_{company_index}')
        
        # Ищем строку с названием компании в листе потребности
        company_row = self._find_company_row(ws, company_name, company_index)
        
        if company_row is None:
            print(f"  Не найдена строка для компании '{company_name}' в листе потребности")
            return
        
        print(f"  Найдена строка {company_row} для компании '{company_name}'")
        
        # Заполняем числовые данные
        data_to_fill = {
            'gasoline_demand': company.get('gasoline_demand', 0),  # Бензин всего (год)
            'diesel_demand': company.get('diesel_demand', 0),      # Дизель всего (год)
            'monthly_gasoline': company.get('monthly_gasoline', 0), # Бензин (месяц)
            'monthly_diesel': company.get('monthly_diesel', 0),     # Дизель (месяц)
        }
        
        # Ищем колонки для заполнения
        for col in range(1, 20):
            try:
                header_cell = ws.cell(row=company_row-1, column=col)  # Предполагаем, что заголовок на строке выше
                if header_cell.value and isinstance(header_cell.value, str):
                    header = str(header_cell.value).lower()
                    
                    # Определяем что это за колонка и заполняем
                    if any(word in header for word in ['бензин', 'gasoline', '92', '95']):
                        if 'год' in header or 'всего' in header:
                            value_cell = ws.cell(row=company_row, column=col)
                            value_cell.value = data_to_fill['gasoline_demand']
                            value_cell.number_format = '0.00'
                            print(f"    Бензин (год): {data_to_fill['gasoline_demand']} в {value_cell.coordinate}")
                        elif 'месяц' in header or 'мес' in header:
                            value_cell = ws.cell(row=company_row, column=col)
                            value_cell.value = data_to_fill['monthly_gasoline']
                            value_cell.number_format = '0.00'
                            print(f"    Бензин (мес): {data_to_fill['monthly_gasoline']} в {value_cell.coordinate}")
                    
                    elif any(word in header for word in ['дизель', 'diesel', 'диз']):
                        if 'год' in header or 'всего' in header:
                            value_cell = ws.cell(row=company_row, column=col)
                            value_cell.value = data_to_fill['diesel_demand']
                            value_cell.number_format = '0.00'
                            print(f"    Дизель (год): {data_to_fill['diesel_demand']} в {value_cell.coordinate}")
                        elif 'месяц' in header or 'мес' in header:
                            value_cell = ws.cell(row=company_row, column=col)
                            value_cell.value = data_to_fill['monthly_diesel']
                            value_cell.number_format = '0.00'
                            print(f"    Дизель (мес): {data_to_fill['monthly_diesel']} в {value_cell.coordinate}")
            except:
                continue
    
    def _fill_company_balance(self, ws, company: Dict[str, Any], company_index: int):
        """Заполнение данных остатков для одной компании"""
        company_name = company.get('name', f'Компания_{company_index}')
        
        # Ищем строку с названием компании
        company_row = self._find_company_row(ws, company_name, company_index)
        
        if company_row is None:
            print(f"  Не найдена строка для компании '{company_name}' в листе остатков")
            return
        
        print(f"  Найдена строка {company_row} для компании '{company_name}' в остатках")
        
        # Заполняем данные остатков
        data_to_fill = {
            'stock_ai92': company.get('stock_ai92', 0),           # АИ-92
            'stock_ai95': company.get('stock_ai95', 0),           # АИ-95
            'stock_diesel_winter': company.get('stock_diesel_winter', 0),  # Дизель зимний
            'stock_diesel_arctic': company.get('stock_diesel_arctic', 0),  # Дизель арктический
        }
        
        # Ищем колонки по заголовкам
        for col in range(1, 30):  # Больше колонок, так как в остатках может быть много данных
            try:
                # Ищем заголовок (обычно на первой строке таблицы)
                header_row = self._find_header_row(ws, company_row)
                if header_row is None:
                    header_row = company_row - 1
                
                header_cell = ws.cell(row=header_row, column=col)
                if header_cell.value and isinstance(header_cell.value, str):
                    header = str(header_cell.value).lower()
                    
                    # Определяем тип топлива по заголовку
                    if '92' in header and any(word in header for word in ['аи', 'ai', 'бензин']):
                        value_cell = ws.cell(row=company_row, column=col)
                        value_cell.value = data_to_fill['stock_ai92']
                        value_cell.number_format = '0.00'
                        print(f"    АИ-92: {data_to_fill['stock_ai92']} в {value_cell.coordinate}")
                    
                    elif '95' in header and any(word in header for word in ['аи', 'ai', 'бензин']):
                        value_cell = ws.cell(row=company_row, column=col)
                        value_cell.value = data_to_fill['stock_ai95']
                        value_cell.number_format = '0.00'
                        print(f"    АИ-95: {data_to_fill['stock_ai95']} в {value_cell.coordinate}")
                    
                    elif 'зим' in header and any(word in header for word in ['дизель', 'диз']):
                        value_cell = ws.cell(row=company_row, column=col)
                        value_cell.value = data_to_fill['stock_diesel_winter']
                        value_cell.number_format = '0.00'
                        print(f"    Дизель зимний: {data_to_fill['stock_diesel_winter']} в {value_cell.coordinate}")
                    
                    elif 'аркт' in header and any(word in header for word in ['дизель', 'диз']):
                        value_cell = ws.cell(row=company_row, column=col)
                        value_cell.value = data_to_fill['stock_diesel_arctic']
                        value_cell.number_format = '0.00'
                        print(f"    Дизель арктический: {data_to_fill['stock_diesel_arctic']} в {value_cell.coordinate}")
            except:
                continue
    
    def _fill_company_supply(self, ws, company: Dict[str, Any], company_index: int):
        """Заполнение данных поставок для одной компании"""
        company_name = company.get('name', f'Компания_{company_index}')
        
        company_row = self._find_company_row(ws, company_name, company_index)
        
        if company_row is None:
            print(f"  Не найдена строка для компании '{company_name}' в листе поставок")
            return
        
        print(f"  Найдена строка {company_row} для компании '{company_name}' в поставках")
        
        # Заполняем данные поставок
        data_to_fill = {
            'supply_ai92': company.get('supply_ai92', 0),
            'supply_ai95': company.get('supply_ai95', 0),
            'supply_diesel_winter': company.get('supply_diesel_winter', 0),
            'supply_diesel_arctic': company.get('supply_diesel_arctic', 0),
        }
        
        # Похожая логика как для остатков
        for col in range(1, 30):
            try:
                header_row = self._find_header_row(ws, company_row)
                if header_row is None:
                    header_row = company_row - 1
                
                header_cell = ws.cell(row=header_row, column=col)
                if header_cell.value and isinstance(header_cell.value, str):
                    header = str(header_cell.value).lower()
                    
                    if '92' in header and any(word in header for word in ['аи', 'ai', 'бензин', 'поставк']):
                        value_cell = ws.cell(row=company_row, column=col)
                        value_cell.value = data_to_fill['supply_ai92']
                        value_cell.number_format = '0.00'
                        print(f"    Поставки АИ-92: {data_to_fill['supply_ai92']} в {value_cell.coordinate}")
                    
                    elif '95' in header and any(word in header for word in ['аи', 'ai', 'бензин', 'поставк']):
                        value_cell = ws.cell(row=company_row, column=col)
                        value_cell.value = data_to_fill['supply_ai95']
                        value_cell.number_format = '0.00'
                        print(f"    Поставки АИ-95: {data_to_fill['supply_ai95']} в {value_cell.coordinate}")
                    
                    elif 'зим' in header and any(word in header for word in ['дизель', 'диз', 'поставк']):
                        value_cell = ws.cell(row=company_row, column=col)
                        value_cell.value = data_to_fill['supply_diesel_winter']
                        value_cell.number_format = '0.00'
                        print(f"    Поставки дизель зимний: {data_to_fill['supply_diesel_winter']} в {value_cell.coordinate}")
                    
                    elif 'аркт' in header and any(word in header for word in ['дизель', 'диз', 'поставк']):
                        value_cell = ws.cell(row=company_row, column=col)
                        value_cell.value = data_to_fill['supply_diesel_arctic']
                        value_cell.number_format = '0.00'
                        print(f"    Поставки дизель арктический: {data_to_fill['supply_diesel_arctic']} в {value_cell.coordinate}")
            except:
                continue
    
    def _fill_company_sales(self, ws, company: Dict[str, Any], company_index: int):
        """Заполнение данных реализации для одной компании"""
        company_name = company.get('name', f'Компания_{company_index}')
        
        company_row = self._find_company_row(ws, company_name, company_index)
        
        if company_row is None:
            print(f"  Не найдена строка для компании '{company_name}' в листе реализации")
            return
        
        print(f"  Найдена строка {company_row} для компании '{company_name}' в реализации")
        
        # Заполняем данные реализации
        data_to_fill = {
            'sales_ai92': company.get('sales_ai92', 0),
            'sales_ai95': company.get('sales_ai95', 0),
            'sales_diesel_winter': company.get('sales_diesel_winter', 0),
            'sales_diesel_arctic': company.get('sales_diesel_arctic', 0),
        }
        
        for col in range(1, 30):
            try:
                header_row = self._find_header_row(ws, company_row)
                if header_row is None:
                    header_row = company_row - 1
                
                header_cell = ws.cell(row=header_row, column=col)
                if header_cell.value and isinstance(header_cell.value, str):
                    header = str(header_cell.value).lower()
                    
                    if '92' in header and any(word in header for word in ['аи', 'ai', 'бензин', 'реализ', 'продаж']):
                        value_cell = ws.cell(row=company_row, column=col)
                        value_cell.value = data_to_fill['sales_ai92']
                        value_cell.number_format = '0.00'
                        print(f"    Реализация АИ-92: {data_to_fill['sales_ai92']} в {value_cell.coordinate}")
                    
                    elif '95' in header and any(word in header for word in ['аи', 'ai', 'бензин', 'реализ', 'продаж']):
                        value_cell = ws.cell(row=company_row, column=col)
                        value_cell.value = data_to_fill['sales_ai95']
                        value_cell.number_format = '0.00'
                        print(f"    Реализация АИ-95: {data_to_fill['sales_ai95']} в {value_cell.coordinate}")
                    
                    elif 'зим' in header and any(word in header for word in ['дизель', 'диз', 'реализ', 'продаж']):
                        value_cell = ws.cell(row=company_row, column=col)
                        value_cell.value = data_to_fill['sales_diesel_winter']
                        value_cell.number_format = '0.00'
                        print(f"    Реализация дизель зимний: {data_to_fill['sales_diesel_winter']} в {value_cell.coordinate}")
                    
                    elif 'аркт' in header and any(word in header for word in ['дизель', 'диз', 'реализ', 'продаж']):
                        value_cell = ws.cell(row=company_row, column=col)
                        value_cell.value = data_to_fill['sales_diesel_arctic']
                        value_cell.number_format = '0.00'
                        print(f"    Реализация дизель арктический: {data_to_fill['sales_diesel_arctic']} в {value_cell.coordinate}")
            except:
                continue
    
    def _find_company_row(self, ws, company_name: str, company_index: int) -> Optional[int]:
        """Находит строку с названием компании"""
        # Сначала ищем по полному названию
        for row in range(1, 200):
            for col in range(1, 10):  # Ищем в первых 10 колонках
                try:
                    cell = ws.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        if company_name.lower() in cell.value.lower():
                            return row
                except:
                    continue
        
        # Если не нашли, ищем по номеру компании
        for row in range(1, 200):
            for col in range(1, 5):  # Номер обычно в первых колонках
                try:
                    cell = ws.cell(row=row, column=col)
                    if cell.value == company_index or str(cell.value) == str(company_index):
                        return row
                except:
                    continue
        
        # Если все еще не нашли, ищем пустые строки в таблице
        table_start = self._find_table_start(ws)
        if table_start:
            # Находим первую пустую строку после начала таблицы
            for row in range(table_start, table_start + 50):
                # Проверяем что строка пустая
                first_cell = ws.cell(row=row, column=1)
                if first_cell.value is None or first_cell.value == '':
                    # Записываем название компании
                    company_cell = ws.cell(row=row, column=2)  # Предполагаем что компания во 2й колонке
                    company_cell.value = company_name
                    print(f"  Добавлена новая строка {row} для компании '{company_name}'")
                    return row
        
        return None
    
    def _find_table_start(self, ws) -> Optional[int]:
        """Находит начало таблицы (первую строку с заголовками)"""
        table_headers = ['№', 'номер', 'компания', 'наименование', '1', '2', '3']
        
        for row in range(1, 100):
            for col in range(1, 10):
                try:
                    cell = ws.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        cell_value = str(cell.value).lower()
                        for header in table_headers:
                            if header in cell_value:
                                return row + 1  # Данные начинаются со следующей строки
                except:
                    continue
        
        return None
    
    def _find_header_row(self, ws, data_row: int) -> Optional[int]:
        """Находит строку с заголовками для данной строки данных"""
        # Ищем заголовки выше строки данных
        for row in range(data_row - 10, data_row):
            if row < 1:
                continue
            
            # Проверяем есть ли в строке заголовки
            for col in range(1, 20):
                try:
                    cell = ws.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        cell_value = str(cell.value).lower()
                        if any(word in cell_value for word in ['аи', 'ai', 'дизель', 'бензин', '92', '95']):
                            return row
                except:
                    continue
        
        return None