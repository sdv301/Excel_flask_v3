# app.py - ПОЛНЫЙ ИСПРАВЛЕННЫЙ КОД
from flask import Flask, render_template, request, jsonify, send_file
import os
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import traceback

from config import Config

from reports.report_generator import SummaryReportGenerator, TemplateReportGenerator, SimpleTemplateReportGenerator

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


from database.connection import db_connection
from database.queries import db
from database.models import (
    Company, 
    UploadedFile, 
    Sheet1Structure, 
    Sheet2Demand, 
    Sheet3Balance, 
    Sheet4Supply, 
    Sheet5Sales,
    Sheet6Aviation,
    Sheet7Comments
)

# Проверяем доступность парсеров
PARSER_AVAILABLE = False
SIMPLE_PARSER_AVAILABLE = False
SIMPLE_ALL_PARSER_AVAILABLE = False
NEW_PARSER_AVAILABLE = False

try:
    from parser.excel_parser import FuelReportParser
    PARSER_AVAILABLE = True
    print("✓ Основной парсер доступен")
except ImportError as e:
    print(f"✗ Основной парсер не доступен: {e}")

try:
    from parser.simple_parser import SimpleFuelParser
    SIMPLE_PARSER_AVAILABLE = True
    print("✓ Простой парсер доступен")
except ImportError as e:
    print(f"✗ Простой парсер не доступен: {e}")

try:
    from parser.simple_all_parser import SimpleAllParser
    SIMPLE_ALL_PARSER_AVAILABLE = True
    print("✓ Упрощенный парсер всех листов доступен")
except ImportError as e:
    print(f"✗ Упрощенный парсер всех листов не доступен: {e}")

try:
    from parser.simple_all_parser_fixed_v2 import SimpleAllParserV2
    NEW_PARSER_AVAILABLE = True
    print("✓ Новый улучшенный парсер доступен")
except ImportError as e:
    print(f"✗ Новый улучшенный парсер не доступен: {e}")

app = Flask(__name__)
app.config.from_object(Config)

# Создаем таблицы в БД при первом запуске
with app.app_context():
    try:
        db_connection.create_tables()
        print("Таблицы базы данных созданы успешно")
        
        # Добавляем тестовые компании если их нет
        session = db_connection.get_session()
        from database.models import Company
        
        existing = session.query(Company).count()
        if existing == 0:
            test_companies = [
                ("Саханефтегазсбыт", "СНГС"),
                ("Туймаада-Нефть", "ТУЙМААДА"),
                ("Сибойл", "СИБОЙЛ"),
                ("ЭКТО-Ойл", "ЭКТО"),
                ("Сибирское топливо", "СИБТОП"),
                ("Паритет", "ПАРИТЕТ")
            ]
            
            for name, code in test_companies:
                company = Company(name=name, code=code)
                session.add(company)
            
            session.commit()
            print("Тестовые компании добавлены")
    except Exception as e:
        print(f"Ошибка при инициализации БД: {e}")
    finally:
        db_connection.close_session()

@app.route('/')
def index():
    """Главная страница"""
    try:
        companies = db.get_companies()
        recent_files = db.get_recent_files(limit=10)
        return render_template('index.html', 
                             companies=companies, 
                             recent_files=recent_files,
                             now=datetime.now())
    except Exception as e:
        return f"Ошибка: {str(e)}<br>{traceback.format_exc()}"

@app.route('/admin')
def admin():
    """Главная страница"""
    try:
        companies = db.get_companies()
        recent_files = db.get_recent_files(limit=10)
        return render_template('admin.html', 
                             companies=companies, 
                             recent_files=recent_files,
                             now=datetime.now())
    except Exception as e:
        return f"Ошибка: {str(e)}<br>{traceback.format_exc()}"

@app.route('/upload', methods=['POST'])
def upload_file():
    """Загрузка файла"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Файл не выбран'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Файл не выбран'}), 400
        
        if not file.filename.lower().endswith('.xlsx'):
            return jsonify({'error': 'Только Excel файлы (.xlsx)'}), 400
        
        # Сохраняем файл
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        print(f"\n=== НАЧАЛО ОБРАБОТКИ ФАЙЛА: {filename} ===")
        print(f"Файл сохранен: {file_path}")
        
        # Сначала пробуем новый улучшенный парсер
        if NEW_PARSER_AVAILABLE:
            print("Пробуем использовать новый улучшенный парсер...")
            return _process_with_new_parser(filename, file_path)
        
        # Затем упрощенный парсер всех листов
        elif SIMPLE_ALL_PARSER_AVAILABLE:
            print("Пробуем использовать упрощенный парсер всех листов...")
            return _process_with_simple_all_parser(filename, file_path)
        
        # Затем основной парсер
        elif PARSER_AVAILABLE:
            print("Пробуем использовать основной парсер...")
            return _process_with_main_parser(filename, file_path)
        
        # И наконец простой парсер
        elif SIMPLE_PARSER_AVAILABLE:
            print("Используется простой парсер...")
            return _process_with_simple_parser(filename, file_path)
        
        else:
            return jsonify({'error': 'Нет доступных парсеров'}), 500
        
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"Ошибка при загрузке файла: {error_details}")
        return jsonify({'error': str(e), 'details': error_details}), 500

def _process_with_new_parser(filename, file_path):
    """Обработка файла новым улучшенным парсером"""
    try:
        parser = SimpleAllParserV2(file_path)
        all_data = parser.parse_all()
        
        metadata = all_data['metadata']
        
        print(f"\nНовый парсер результаты:")
        print(f"  Компания: {metadata['company']}")
        print(f"  Лист 1: {len(all_data.get('sheet1', []))} записей")
        print(f"  Лист 2: {len(all_data.get('sheet2', []))} записей")
        print(f"  Лист 3: {len(all_data.get('sheet3', []))} записей")
        print(f"  Лист 4: {len(all_data.get('sheet4', []))} записей")
        print(f"  Лист 5: {len(all_data.get('sheet5', []))} записей")
        
        # Сохраняем в БД информацию о файле
        file_id, company_id = db.save_uploaded_file(
            filename=filename,
            file_path=file_path,
            company_name=metadata['company'],
            report_date=metadata['report_date'].date()
        )
        
        print(f"Файл сохранен в БД: ID={file_id}, Company ID={company_id}")
        
        # Сохраняем все данные
        saved_counts = {}
        
        # Лист 2 (Потребность) - сохраняем как одну запись
        if all_data.get('sheet2'):
            try:
                # Берем первую запись (месячную потребность)
                if all_data['sheet2']:
                    demand_data = all_data['sheet2'][0]
                    db.save_sheet2_data(file_id, company_id, metadata['report_date'].date(), demand_data)
                    saved_counts['sheet2'] = 1
                    print(f"✓ Лист 2 сохранен: месячная потребность")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 2: {e}")
                saved_counts['sheet2'] = 0
        
        # Лист 3 (Остатки)
        if all_data.get('sheet3'):
            try:
                db.save_sheet3_data(file_id, company_id, metadata['report_date'].date(), all_data['sheet3'])
                saved_counts['sheet3'] = len(all_data['sheet3'])
                print(f"✓ Лист 3 сохранен: {len(all_data['sheet3'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 3: {e}")
                saved_counts['sheet3'] = 0
        
        # Лист 4 (Поставки)
        if all_data.get('sheet4'):
            try:
                db.save_sheet4_data(file_id, company_id, metadata['report_date'].date(), all_data['sheet4'])
                saved_counts['sheet4'] = len(all_data['sheet4'])
                print(f"✓ Лист 4 сохранен: {len(all_data['sheet4'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 4: {e}")
                saved_counts['sheet4'] = 0
        
        # Лист 5 (Реализация)
        if all_data.get('sheet5'):
            try:
                db.save_sheet5_data(file_id, company_id, metadata['report_date'].date(), all_data['sheet5'])
                saved_counts['sheet5'] = len(all_data['sheet5'])
                print(f"✓ Лист 5 сохранен: {len(all_data['sheet5'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 5: {e}")
                saved_counts['sheet5'] = 0

        # Обновляем статус файла в БД
        try:
            db.update_file_status(file_id, 'processed')
            print(f"✓ Статус файла обновлен на 'processed'")
        except Exception as e:
            print(f"⚠ Не удалось обновить статус файла: {e}")
        
        print(f"=== ЗАВЕРШЕНО ОБРАБОТКА ФАЙЛА: {filename} ===\n")
        
        return jsonify({
            'success': True,
            'message': 'Файл успешно обработан (новый улучшенный парсер)',
            'company': metadata['company'],
            'report_date': metadata['report_date'].strftime('%Y-%m-%d'),
            'data_extracted': {
                'sheet1': len(all_data.get('sheet1', [])),
                'sheet2': len(all_data.get('sheet2', [])),
                'sheet3': len(all_data.get('sheet3', [])),
                'sheet4': len(all_data.get('sheet4', [])),
                'sheet5': len(all_data.get('sheet5', [])),
            },
            'data_saved': saved_counts,
            'file_info': {
                'file_id': file_id,
                'company_id': company_id,
                'filename': filename
            }
        })
        
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"Ошибка нового парсера: {error_details}")
        
        # Пробуем следующий доступный парсер
        if SIMPLE_ALL_PARSER_AVAILABLE:
            print("Пробуем использовать упрощенный парсер всех листов...")
            return _process_with_simple_all_parser(filename, file_path)
        elif PARSER_AVAILABLE:
            print("Пробуем использовать основной парсер...")
            return _process_with_main_parser(filename, file_path)
        elif SIMPLE_PARSER_AVAILABLE:
            print("Пробуем использовать простой парсер...")
            return _process_with_simple_parser(filename, file_path)
        else:
            return jsonify({
                'error': f'Ошибка парсинга: {str(e)}',
                'details': error_details
            }), 500
            
def _process_with_simple_all_parser(filename, file_path):
    """Обработка файла упрощенным парсером всех листов"""
    try:
        parser = SimpleAllParser(file_path)
        all_data = parser.parse_all()
        
        metadata = all_data['metadata']
        
        print(f"Упрощенный парсер всех листов:")
        print(f"  Компания: {metadata['company']}")
        print(f"  Дата: {metadata['report_date']}")
        print(f"  Лист 1: {len(all_data.get('sheet1', []))} записей")
        print(f"  Лист 2: {'есть' if all_data.get('sheet2') else 'нет'}")
        print(f"  Лист 3: {len(all_data.get('sheet3', []))} записей")
        print(f"  Лист 4: {len(all_data.get('sheet4', []))} записей")
        print(f"  Лист 5: {len(all_data.get('sheet5', []))} записей")
        
        # Сохраняем в БД
        file_id, company_id = db.save_uploaded_file(
            filename=filename,
            file_path=file_path,
            company_name=metadata['company'],
            report_date=metadata['report_date'].date()
        )
        
        # Сохраняем все данные
        saved_counts = {}
        
        if all_data.get('sheet1'):
            try:
                db.save_sheet1_data(file_id, company_id, metadata['report_date'].date(), all_data['sheet1'])
                saved_counts['sheet1'] = len(all_data['sheet1'])
                print(f"✓ Лист 1 сохранен: {len(all_data['sheet1'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 1: {e}")
                saved_counts['sheet1'] = 0
        
        if all_data.get('sheet2'):
            try:
                db.save_sheet2_data(file_id, company_id, metadata['report_date'].date(), all_data['sheet2'])
                saved_counts['sheet2'] = 1
                print(f"✓ Лист 2 сохранен")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 2: {e}")
                saved_counts['sheet2'] = 0
        
        if all_data.get('sheet3'):
            try:
                db.save_sheet3_data(file_id, company_id, metadata['report_date'].date(), all_data['sheet3'])
                saved_counts['sheet3'] = len(all_data['sheet3'])
                print(f"✓ Лист 3 сохранен: {len(all_data['sheet3'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 3: {e}")
                saved_counts['sheet3'] = 0
        
        if all_data.get('sheet4'):
            try:
                db.save_sheet4_data(file_id, company_id, metadata['report_date'].date(), all_data['sheet4'])
                saved_counts['sheet4'] = len(all_data['sheet4'])
                print(f"✓ Лист 4 сохранен: {len(all_data['sheet4'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 4: {e}")
                saved_counts['sheet4'] = 0
        
        if all_data.get('sheet5'):
            try:
                db.save_sheet5_data(file_id, company_id, metadata['report_date'].date(), all_data['sheet5'])
                saved_counts['sheet5'] = len(all_data['sheet5'])
                print(f"✓ Лист 5 сохранен: {len(all_data['sheet5'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 5: {e}")
                saved_counts['sheet5'] = 0
        
        # Обновляем статус
        db.update_file_status(file_id, 'processed')
        
        print(f"=== ЗАВЕРШЕНО ОБРАБОТКА ФАЙЛА: {filename} ===\n")
        
        return jsonify({
            'success': True,
            'message': 'Файл успешно обработан (упрощенный парсер всех листов)',
            'company': metadata['company'],
            'report_date': metadata['report_date'].strftime('%Y-%m-%d'),
            'data_extracted': {
                'sheet1': len(all_data.get('sheet1', [])),
                'sheet2': 1 if all_data.get('sheet2') else 0,
                'sheet3': len(all_data.get('sheet3', [])),
                'sheet4': len(all_data.get('sheet4', [])),
                'sheet5': len(all_data.get('sheet5', []))
            },
            'data_saved': saved_counts
        })
        
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"Ошибка упрощенного парсера всех листов: {error_details}")
        
        # Пробуем простой парсер
        if SIMPLE_PARSER_AVAILABLE:
            print("Пробуем использовать простой парсер...")
            return _process_with_simple_parser(filename, file_path)
        else:
            return jsonify({'error': str(e), 'details': error_details}), 500

def _process_with_main_parser(filename, file_path):
    """Обработка файла основным парсером"""
    try:
        parser = FuelReportParser(file_path)
        metadata = parser.parse()
        
        print(f"Основной парсер: компания={metadata.company_name}, дата={metadata.report_date}")
        
        # Сохраняем в БД информацию о файле
        file_id, company_id = db.save_uploaded_file(
            filename=filename,
            file_path=file_path,
            company_name=metadata.company_name,
            report_date=metadata.report_date.date()
        )
        
        print(f"Файл сохранен в БД: ID={file_id}, Company ID={company_id}")
        
        # Извлекаем ВСЕ данные из файла
        print("Извлечение данных из файла...")
        all_data = parser.extract_all_data()
        
        print(f"Извлечено данных:")
        print(f"  Лист 1: {len(all_data.get('sheet1', []))} записей")
        print(f"  Лист 2: {'есть' if all_data.get('sheet2') else 'нет'}")
        print(f"  Лист 3: {len(all_data.get('sheet3', []))} записей")
        print(f"  Лист 4: {len(all_data.get('sheet4', []))} записей")
        print(f"  Лист 5: {len(all_data.get('sheet5', []))} записей")
        print(f"  Лист 6: {len(all_data.get('sheet6', []))} записей")
        print(f"  Лист 7: {len(all_data.get('sheet7', []))} записей")
        
        # Сохраняем данные из ВСЕХ листов
        saved_counts = {}
        
        # Лист 1
        if all_data.get('sheet1'):
            try:
                db.save_sheet1_data(file_id, company_id, metadata.report_date.date(), all_data['sheet1'])
                saved_counts['sheet1'] = len(all_data['sheet1'])
                print(f"✓ Лист 1 сохранен: {len(all_data['sheet1'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 1: {e}")
                saved_counts['sheet1'] = 0
        
        # Лист 2
        if all_data.get('sheet2'):
            try:
                db.save_sheet2_data(file_id, company_id, metadata.report_date.date(), all_data['sheet2'])
                saved_counts['sheet2'] = 1
                print(f"✓ Лист 2 сохранен")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 2: {e}")
                saved_counts['sheet2'] = 0
        
        # Лист 3
        if all_data.get('sheet3'):
            try:
                db.save_sheet3_data(file_id, company_id, metadata.report_date.date(), all_data['sheet3'])
                saved_counts['sheet3'] = len(all_data['sheet3'])
                print(f"✓ Лист 3 сохранен: {len(all_data['sheet3'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 3: {e}")
                saved_counts['sheet3'] = 0
        
        # Лист 4
        if all_data.get('sheet4'):
            try:
                db.save_sheet4_data(file_id, company_id, metadata.report_date.date(), all_data['sheet4'])
                saved_counts['sheet4'] = len(all_data['sheet4'])
                print(f"✓ Лист 4 сохранен: {len(all_data['sheet4'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 4: {e}")
                saved_counts['sheet4'] = 0
        
        # Лист 5
        if all_data.get('sheet5'):
            try:
                db.save_sheet5_data(file_id, company_id, metadata.report_date.date(), all_data['sheet5'])
                saved_counts['sheet5'] = len(all_data['sheet5'])
                print(f"✓ Лист 5 сохранен: {len(all_data['sheet5'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 5: {e}")
                saved_counts['sheet5'] = 0
        
        # Лист 6
        if all_data.get('sheet6'):
            try:
                db.save_sheet6_data(file_id, company_id, metadata.report_date.date(), all_data['sheet6'])
                saved_counts['sheet6'] = len(all_data['sheet6'])
                print(f"✓ Лист 6 сохранен: {len(all_data['sheet6'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 6: {e}")
                saved_counts['sheet6'] = 0
        
        # Лист 7
        if all_data.get('sheet7'):
            try:
                db.save_sheet7_data(file_id, company_id, metadata.report_date.date(), all_data['sheet7'])
                saved_counts['sheet7'] = len(all_data['sheet7'])
                print(f"✓ Лист 7 сохранен: {len(all_data['sheet7'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 7: {e}")
                saved_counts['sheet7'] = 0
        
        # Обновляем статус файла в БД
        try:
            db.update_file_status(file_id, 'processed')
            print(f"✓ Статус файла обновлен на 'processed'")
        except Exception as e:
            print(f"⚠ Не удалось обновить статус файла: {e}")
        
        print(f"=== ЗАВЕРШЕНО ОБРАБОТКА ФАЙЛА: {filename} ===\n")
        
        return jsonify({
            'success': True,
            'message': 'Файл успешно обработан (основной парсер)',
            'company': metadata.company_name,
            'report_date': metadata.report_date.strftime('%Y-%m-%d'),
            'data_extracted': {
                'sheet1': len(all_data.get('sheet1', [])),
                'sheet2': 1 if all_data.get('sheet2') else 0,
                'sheet3': len(all_data.get('sheet3', [])),
                'sheet4': len(all_data.get('sheet4', [])),
                'sheet5': len(all_data.get('sheet5', [])),
                'sheet6': len(all_data.get('sheet6', [])),
                'sheet7': len(all_data.get('sheet7', []))
            },
            'data_saved': saved_counts,
            'file_info': {
                'file_id': file_id,
                'company_id': company_id,
                'filename': filename
            }
        })
        
    except Exception as parse_error:
        print(f"Ошибка при обработке основным парсером: {parse_error}")
        traceback.print_exc()
        
        # Пробуем следующий доступный парсер
        if SIMPLE_ALL_PARSER_AVAILABLE:
            print("Пробуем использовать упрощенный парсер всех листов...")
            return _process_with_simple_all_parser(filename, file_path)
        elif SIMPLE_PARSER_AVAILABLE:
            print("Пробуем использовать простой парсер...")
            return _process_with_simple_parser(filename, file_path)
        else:
            return jsonify({
                'error': f'Ошибка парсинга: {str(parse_error)}',
                'details': traceback.format_exc()
            }), 500

def _process_with_simple_parser(filename, file_path):
    """Обработка файла простым парсером"""
    try:
        parser = SimpleFuelParser(file_path)
        data = parser.parse_all()
        
        print(f"Простой парсер: компания={data.get('company', 'Неизвестно')}")
        print(f"  Лист 1: {len(data.get('sheet1', []))} записей")
        print(f"  Лист 2: {'есть' if data.get('sheet2') else 'нет'}")
        print(f"  Лист 3: {len(data.get('sheet3', []))} записей")
        
        # Сохраняем в БД
        file_id, company_id = db.save_uploaded_file(
            filename=filename,
            file_path=file_path,
            company_name=data.get('company', 'Неизвестная компания'),
            report_date=datetime.now().date()
        )
        
        saved_counts = {}
        
        # Сохраняем данные
        if data.get('sheet1'):
            try:
                db.save_sheet1_data(file_id, company_id, datetime.now().date(), data['sheet1'])
                saved_counts['sheet1'] = len(data['sheet1'])
                print(f"✓ Лист 1 сохранен: {len(data['sheet1'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 1: {e}")
                saved_counts['sheet1'] = 0
        
        if data.get('sheet2'):
            try:
                db.save_sheet2_data(file_id, company_id, datetime.now().date(), data['sheet2'])
                saved_counts['sheet2'] = 1
                print(f"✓ Лист 2 сохранен")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 2: {e}")
                saved_counts['sheet2'] = 0
        
        if data.get('sheet3'):
            try:
                db.save_sheet3_data(file_id, company_id, datetime.now().date(), data['sheet3'])
                saved_counts['sheet3'] = len(data['sheet3'])
                print(f"✓ Лист 3 сохранен: {len(data['sheet3'])} записей")
            except Exception as e:
                print(f"✗ Ошибка сохранения листа 3: {e}")
                saved_counts['sheet3'] = 0
        
        # Обновляем статус
        db.update_file_status(file_id, 'processed')
        
        print(f"=== ЗАВЕРШЕНО ОБРАБОТКА ФАЙЛА: {filename} ===\n")
        
        return jsonify({
            'success': True,
            'message': 'Файл успешно обработан (простой парсер)',
            'company': data.get('company', 'Неизвестная компания'),
            'report_date': datetime.now().strftime('%Y-%m-%d'),
            'data_extracted': {
                'sheet1': len(data.get('sheet1', [])),
                'sheet2': 1 if data.get('sheet2') else 0,
                'sheet3': len(data.get('sheet3', []))
            },
            'data_saved': saved_counts
        })
        
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"Ошибка простого парсера: {error_details}")
        return jsonify({'error': str(e), 'details': error_details}), 500

@app.route('/generate-report', methods=['GET', 'POST'])
def generate_report():
    """Генерация сводного отчета (работает с GET и POST)"""
    try:
        # Определяем дату отчета
        report_date = None
        
        if request.method == 'POST':
            if request.is_json:
                data = request.get_json()
                report_date_str = data.get('report_date')
            else:
                report_date_str = request.form.get('report_date')
        else:
            report_date_str = request.args.get('report_date')
        
        # Устанавливаем дату
        if not report_date_str:
            report_date = datetime.now().date()
        else:
            report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
        
        print(f"\n=== ГЕНЕРАЦИЯ ОТЧЕТА ===")
        print(f"Метод запроса: {request.method}")
        print(f"Запрошена дата: {report_date}")
        
        # Создаем генератор отчетов
        generator = SummaryReportGenerator(db)
        
        # Ищем данные БЕЗ фильтра по дате - метод сам найдет последние данные
        print("Ищем последние данные без фильтра по дате...")
        aggregated_data = db.get_aggregated_data()  # Без аргументов = все последние данные
        
        if not aggregated_data:
            # Если не нашли, пробуем найти любые данные
            print("Не найдены последние данные, ищем любые данные...")
            session = db_connection.get_session()
            
            # Проверяем есть ли вообще данные в базе
            has_sheet3 = session.query(Sheet3Balance).count() > 0
            has_sheet5 = session.query(Sheet5Sales).count() > 0
            
            if has_sheet3 or has_sheet5:
                # Пробуем получить данные без фильтра по дате
                aggregated_data = db.get_aggregated_data(None)  # Явно передаем None
            else:
                aggregated_data = None
            
            db_connection.close_session()
        
        if not aggregated_data:
            if request.method == 'GET':
                return "<h1>Ошибка</h1><p>Нет данных для генерации отчета</p>", 400
            else:
                return jsonify({'error': 'Нет данных для генерации отчета'}), 400
        
        print(f"Найдено компаний для отчета: {len(aggregated_data)}")
        print("Список компаний:")
        for company_name in aggregated_data.keys():
            company_data = aggregated_data[company_name]
            sheet3 = company_data.get('sheet3_totals', {})
            sheet5 = company_data.get('sheet5_totals', {})
            print(f"  - {company_name}: ")
            print(f"      Остатки AI92: {sheet3.get('total_stock_ai92', 0):.3f}, AI95: {sheet3.get('total_stock_ai95', 0):.3f}")
            print(f"      Реализация AI92: {sheet5.get('total_monthly_ai92', 0):.3f}, AI95: {sheet5.get('total_monthly_ai95', 0):.3f}")
        
        # Генерируем отчет
        report_path = generator.generate_summary_report_with_data(aggregated_data, report_date)
        report_filename = os.path.basename(report_path)
        
        # В зависимости от типа запроса возвращаем разные ответы
        if request.method == 'GET':
            # Для GET запросов возвращаем HTML страницу
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Отчет сгенерирован</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; }}
                    .success {{ color: green; font-size: 18px; }}
                    .info {{ margin: 20px 0; padding: 15px; background: #f0f0f0; border-radius: 5px; }}
                    .btn {{
                        display: inline-block;
                        padding: 12px 24px;
                        background: #2196F3;
                        color: white;
                        text-decoration: none;
                        border-radius: 5px;
                        margin: 10px 0;
                        font-weight: bold;
                        font-size: 16px;
                    }}
                    .btn:hover {{ background: #1976D2; }}
                    .btn-download {{
                        background: #4CAF50;
                        font-size: 18px;
                        padding: 15px 30px;
                    }}
                    .btn-download:hover {{ background: #45a049; }}
                    .date-info {{ color: #666; font-style: italic; }}
                    .company-list {{ max-height: 200px; overflow-y: auto; border: 1px solid #ddd; padding: 10px; }}
                </style>
            </head>
            <body>
                <h1>✅ Отчет успешно сгенерирован!</h1>
                
                <div class="info">
                    <p><strong>📄 Файл:</strong> {report_filename}</p>
                    <p><strong>🏢 Компаний в отчете:</strong> {len(aggregated_data)}</p>
                    <p><strong>📅 Дата генерации:</strong> {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
                    <p class="date-info">Запрошена дата: {report_date.strftime('%d.%m.%Y')}</p>
                </div>
                
                <a class="btn btn-download" href="/download-report/{report_filename}">📥 Скачать отчет Excel</a>
                
                <br><br>
                <div style="margin-top: 30px;">
                    <a class="btn" href="/">← На главную</a> | 
                    <a class="btn" href="/test-generate">Тестовая генерация</a> | 
                    <a class="btn" href="/debug-data">Отладка данных</a>
                </div>
                
                <div style="margin-top: 20px; font-size: 12px; color: #888;">
                    <p>Список компаний в отчете:</p>
                    <div class="company-list">
                        <ul>
            """
            
            # Добавляем список компаний с данными
            for company_name, company_data in aggregated_data.items():
                sheet3 = company_data.get('sheet3_totals', {})
                sheet5 = company_data.get('sheet5_totals', {})
                html += f"""
                        <li>
                            <strong>{company_name}</strong>
                            <br>Остатки: AI92={sheet3.get('total_stock_ai92', 0):.3f}т, AI95={sheet3.get('total_stock_ai95', 0):.3f}т
                            <br>Реализация: AI92={sheet5.get('total_monthly_ai92', 0):.3f}т/мес
                        </li>
                """
            
            html += """
                        </ul>
                    </div>
                </div>
            </body>
            </html>
            """
            
            return html
        else:
            # Для POST запросов возвращаем JSON
            return jsonify({
                'success': True,
                'report_path': report_path,
                'filename': report_filename,
                'download_url': f'/download-report/{report_filename}',
                'message': f'Отчет успешно сгенерирован: {report_filename}',
                'details': {
                    'companies_count': len(aggregated_data),
                    'generation_date': datetime.now().strftime('%Y-%m-%d %H:%M'),
                    'requested_date': report_date.strftime('%Y-%m-%d'),
                    'companies': list(aggregated_data.keys())
                }
            })
        
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"Ошибка при генерации отчета: {error_details}")
        
        if request.method == 'GET':
            html = f"""
            <!DOCTYPE html>
            <html>
            <body>
                <h1>❌ Ошибка генерации отчета</h1>
                <p style="color: red;">{str(e)}</p>
                <pre style="background: #f5f5f5; padding: 10px; overflow: auto;">{error_details}</pre>
                <a href="/">← На главную</a>
            </body>
            </html>
            """
            return html, 500
        else:
            return jsonify({'error': str(e), 'details': error_details}), 500

@app.route('/generate-template-report', methods=['GET', 'POST'])
def generate_template_report():
    """Генерация отчета по шаблону (простая версия)"""
    try:
        # Получаем дату отчета
        if request.method == 'POST':
            if request.is_json:
                data = request.get_json()
                report_date_str = data.get('report_date')
            else:
                report_date_str = request.form.get('report_date')
        else:
            report_date_str = request.args.get('report_date')
        
        if not report_date_str:
            report_date = datetime.now().date()
        else:
            report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
        
        print(f"\n=== ГЕНЕРАЦИЯ ОТЧЕТА ПО ШАБЛОНУ (ПРОСТАЯ ВЕРСИЯ) ===")
        print(f"Дата: {report_date}")
        
        # Проверяем существование папки для шаблонов
        template_dir = 'report_templates'
        if not os.path.exists(template_dir):
            os.makedirs(template_dir)
            print(f"Создана папка для шаблонов: {template_dir}")
        
        # Ищем шаблон
        template_path = os.path.join(template_dir, 'Сводный_отчет_шаблон.xlsx')
        
        if not os.path.exists(template_path):
            return "Шаблон не найден. Поместите файл 'Сводный_отчет_шаблон.xlsx' в папку report_templates"
        
        # Создаем простой генератор
        template_generator = SimpleTemplateReportGenerator(db, template_path)
        
        # Генерируем отчет
        report_path = template_generator.generate_from_template(report_date)
        report_filename = os.path.basename(report_path)
        
        # Для GET запросов возвращаем HTML
        if request.method == 'GET':
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Отчет по шаблону</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; }}
                    .success {{ color: green; font-size: 18px; }}
                    .info {{ margin: 20px 0; padding: 15px; background: #f0f8ff; border-radius: 5px; }}
                    .btn {{
                        display: inline-block;
                        padding: 12px 24px;
                        background: #2196F3;
                        color: white;
                        text-decoration: none;
                        border-radius: 5px;
                        margin: 10px 0;
                        font-weight: bold;
                        font-size: 16px;
                    }}
                    .btn:hover {{ background: #1976D2; }}
                    .btn-download {{
                        background: #4CAF50;
                        font-size: 18px;
                        padding: 15px 30px;
                    }}
                    .btn-download:hover {{ background: #45a049; }}
                </style>
            </head>
            <body>
                <h1>✅ Отчет заполнен успешно!</h1>
                
                <div class="info">
                    <p><strong>📄 Файл:</strong> {report_filename}</p>
                    <p><strong>📅 Дата отчета:</strong> {report_date.strftime('%d.%m.%Y')}</p>
                    <p><strong>🔧 Тип заполнения:</strong> Только числовые данные в существующие ячейки</p>
                    <p><em>Отчет является точной копией шаблона с заполненными цифровыми данными</em></p>
                </div>
                
                <a class="btn btn-download" href="/download-report/{report_filename}">
                    ⬇️ Скачать заполненный отчет
                </a>
                
                <br><br>
                <div style="margin-top: 30px;">
                    <a class="btn" href="/">← На главную</a>
                    <a class="btn" href="/generate-report">Стандартный отчет</a>
                </div>
            </body>
            </html>
            """
            return html
        else:
            return jsonify({
                'success': True,
                'report_path': report_path,
                'filename': report_filename,
                'download_url': f'/download-report/{report_filename}',
                'message': 'Отчет по шаблону успешно заполнен'
            })
        
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"Ошибка генерации отчета по шаблону: {error_details}")
        
        if request.method == 'GET':
            html = f"""
            <!DOCTYPE html>
            <html>
            <body>
                <h1>❌ Ошибка заполнения отчета</h1>
                <p style="color: red;">{str(e)}</p>
                <p>Возможные причины:</p>
                <ul>
                    <li>Шаблон не содержит нужных листов (2-Потребность, 3-Остатки, 4-Поставка, 5-Реализация)</li>
                    <li>В шаблоне нет строк с названиями компаний</li>
                    <li>Нет данных в базе для заполнения</li>
                </ul>
                <details>
                    <summary>Подробности ошибки</summary>
                    <pre style="background: #f5f5f5; padding: 10px; overflow: auto;">{error_details}</pre>
                </details>
                <br>
                <a href="/">← На главную</a>
            </body>
            </html>
            """
            return html, 500


def create_simple_report_fallback():
    """Создание простого отчета как запасной вариант"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        print("Создаем запасной отчет...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводный отчет"
        
        ws['A1'] = "СВОДНЫЙ ОТЧЕТ ПО ДАННЫМ ИЗ БАЗЫ ДАННЫХ"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Дата генерации: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws.merge_cells('A2:E2')
        
        headers = ["№", "Компания", "АЗС", "Остатки АИ-92", "Остатки АИ-95"]
        row = 4
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
        
        row += 1
        
        # Получаем данные напрямую из БД
        session = db_connection.get_session()
        
        # Ищем все компании с данными
        companies = session.query(Company).all()
        
        idx = 1
        for company in companies:
            # Проверяем есть ли данные
            sheet1_count = session.query(Sheet1Structure).filter(
                Sheet1Structure.company_id == company.id
            ).count()
            
            sheet3_count = session.query(Sheet3Balance).filter(
                Sheet3Balance.company_id == company.id
            ).count()
            
            if sheet1_count > 0 or sheet3_count > 0:
                # АЗС
                azs_items = session.query(Sheet1Structure).filter(
                    Sheet1Structure.company_id == company.id
                ).all()
                total_azs = sum(item.azs_count or 0 for item in azs_items)
                
                # Остатки
                stock_items = session.query(Sheet3Balance).filter(
                    Sheet3Balance.company_id == company.id
                ).all()
                total_ai92 = sum(item.stock_ai92 or 0 for item in stock_items)
                total_ai95 = sum(item.stock_ai95 or 0 for item in stock_items)
                
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=company.name)
                ws.cell(row=row, column=3, value=total_azs)
                ws.cell(row=row, column=4, value=total_ai92)
                ws.cell(row=row, column=5, value=total_ai95)
                
                idx += 1
                row += 1
        
        db_connection.close_session()
        
        if idx == 1:
            ws['A4'] = "Нет данных в базе данных"
            ws.merge_cells('A4:E4')
        
        # Сохраняем
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = os.path.join('reports_output', f'Запасной_отчет_{timestamp}.xlsx')
        
        wb.save(output_path)
        
        print(f"Запасной отчет создан: {output_path}")
        return output_path
        
    except Exception as e:
        print(f"Ошибка создания запасного отчета: {e}")
        raise

@app.route('/test-direct-report')
def test_direct_report():
    """Тестовая генерация отчета напрямую из базы"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        
        print("\n=== ТЕСТ ПРЯМОЙ ГЕНЕРАЦИИ ОТЧЕТА ===")
        
        # Создаем книгу Excel напрямую
        wb = Workbook()
        ws = wb.active
        ws.title = "Тестовый отчет"
        
        # Заголовок
        ws['A1'] = "ТЕСТОВЫЙ ОТЧЕТ ПО ДАННЫМ ИЗ БАЗЫ ДАННЫХ"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = f"Дата генерации: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws.merge_cells('A2:G2')
        
        # Получаем данные напрямую из БД
        session = db_connection.get_session()
        
        # Заголовки таблицы
        headers = ["№", "Компания", "Остатки АИ-92", "Остатки АИ-95", 
                  "Поставки АИ-92", "Реализация АИ-92", "Дата данных"]
        
        row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Получаем все компании с данными
        companies = session.query(Company).all()
        
        idx = 1
        for company in companies:
            print(f"\nПроверяем компанию: {company.name}")
            
            # Получаем последние данные по остаткам
            sheet3_data = session.query(Sheet3Balance).filter(
                Sheet3Balance.company_id == company.id
            ).order_by(Sheet3Balance.report_date.desc()).first()
            
            # Получаем последние данные по реализации
            sheet5_data = session.query(Sheet5Sales).filter(
                Sheet5Sales.company_id == company.id
            ).order_by(Sheet5Sales.report_date.desc()).first()
            
            # Получаем последние данные по поставкам
            sheet4_data = session.query(Sheet4Supply).filter(
                Sheet4Supply.company_id == company.id
            ).order_by(Sheet4Supply.report_date.desc()).first()
            
            # Если есть хоть какие-то данные
            if sheet3_data or sheet5_data or sheet4_data:
                stock_ai92 = sheet3_data.stock_ai92 if sheet3_data else 0
                stock_ai95 = sheet3_data.stock_ai95 if sheet3_data else 0
                sales_ai92 = sheet5_data.monthly_ai92 if sheet5_data else 0
                supply_ai92 = sheet4_data.supply_ai92 if sheet4_data else 0
                data_date = sheet3_data.report_date if sheet3_data else (sheet5_data.report_date if sheet5_data else None)
                
                print(f"  Данные найдены: AI92={stock_ai92}, AI95={stock_ai95}")
                
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=company.name)
                ws.cell(row=row, column=3, value=float(stock_ai92 or 0))
                ws.cell(row=row, column=4, value=float(stock_ai95 or 0))
                ws.cell(row=row, column=5, value=float(supply_ai92 or 0))
                ws.cell(row=row, column=6, value=float(sales_ai92 or 0))
                ws.cell(row=row, column=7, value=data_date.strftime('%d.%m.%Y') if data_date else '')
                
                # Форматирование числовых ячеек
                for col in [3, 4, 5, 6]:
                    ws.cell(row=row, column=col).number_format = '0.000'
                
                idx += 1
                row += 1
        
        db_connection.close_session()
        
        if idx == 1:
            ws.cell(row=row, column=1, value="Нет данных в базе данных")
            ws.merge_cells(f'A{row}:G{row}')
        else:
            # Итоговая строка
            ws.cell(row=row, column=1, value="ИТОГО:")
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            # Формулы для итогов
            for col in range(3, 7):
                ws.cell(row=row, column=col, value=f"=SUM({chr(64+col)}5:{chr(64+col)}{row-1})")
                ws.cell(row=row, column=col).font = Font(bold=True)
        
        # Автоподбор ширины
        from openpyxl.utils import get_column_letter
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = os.path.join('reports_output', f'Тестовый_прямой_отчет_{timestamp}.xlsx')
        
        wb.save(output_path)
        
        # Возвращаем файл для скачивания
        return send_file(
            output_path,
            as_attachment=True,
            download_name=f'Тестовый_отчет_{timestamp}.xlsx'
        )
        
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"Ошибка тестовой генерации: {error_details}")
        return f"Ошибка: {str(e)}<br><pre>{error_details}</pre>"

@app.route('/check-db-data')
def check_db_data():
    """Проверка данных в базе"""
    try:
        session = db_connection.get_session()
        
        html = "<h1>Проверка данных в базе</h1>"
        
        # 1. Проверяем компании
        html += "<h2>Компании:</h2><ul>"
        companies = session.query(Company).all()
        for company in companies:
            html += f"<li>{company.name} (ID: {company.id}, Активна: {company.is_active})</li>"
        html += "</ul>"
        
        # 2. Проверяем данные Sheet3Balance
        html += "<h2>Данные Sheet3Balance (остатки):</h2><table border='1'>"
        html += "<tr><th>ID</th><th>Компания</th><th>AI92</th><th>AI95</th><th>Дата</th><th>Company ID</th></tr>"
        
        balances = session.query(Sheet3Balance).all()
        for balance in balances:
            html += f"<tr><td>{balance.id}</td><td>{balance.company_name}</td>"
            html += f"<td>{balance.stock_ai92}</td><td>{balance.stock_ai95}</td>"
            html += f"<td>{balance.report_date}</td><td>{balance.company_id}</td></tr>"
        html += "</table>"
        
        # 3. Проверяем связь компаний с данными
        html += "<h2>Связь компаний с данными:</h2><table border='1'>"
        html += "<tr><th>Компания</th><th>Sheet3 записей</th><th>Sheet5 записей</th></tr>"
        
        for company in companies:
            sheet3_count = session.query(Sheet3Balance).filter(
                Sheet3Balance.company_id == company.id
            ).count()
            
            sheet5_count = session.query(Sheet5Sales).filter(
                Sheet5Sales.company_id == company.id
            ).count()
            
            html += f"<tr><td>{company.name}</td><td>{sheet3_count}</td><td>{sheet5_count}</td></tr>"
        
        html += "</table>"
        
        db_connection.close_session()
        
        html += f"<p><a href='/test-direct-report'>Создать тестовый отчет</a></p>"
        html += f"<p><a href='/'>На главную</a></p>"
        
        return html
        
    except Exception as e:
        return f"Ошибка: {str(e)}<br>{traceback.format_exc()}"

@app.route('/download-report/<filename>')
def download_report(filename):
    """Скачивание отчета"""
    try:
        # Пробуем несколько возможных путей
        possible_paths = [
            os.path.join(app.config['REPORTS_FOLDER'], filename),
            os.path.join('reports_output', filename),
            os.path.join(os.getcwd(), 'reports_output', filename)
        ]
        
        report_path = None
        for path in possible_paths:
            if os.path.exists(path):
                report_path = path
                break
        
        if not report_path:
            return jsonify({'error': f'Файл {filename} не найден'}), 404
        
        print(f"Отчет найден: {report_path}")
        
        return send_file(
            report_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"Ошибка при скачивании отчета: {error_details}")
        return jsonify({'error': str(e), 'details': error_details}), 500

@app.route('/generate-template-fixed', methods=['GET', 'POST'])
def generate_template_fixed():
    """Генерация отчета по фиксированному шаблону"""
    try:
        # Получаем дату отчета
        report_date = None
        
        if request.method == 'POST':
            if request.is_json:
                data = request.get_json()
                report_date_str = data.get('report_date')
            else:
                report_date_str = request.form.get('report_date')
        else:
            report_date_str = request.args.get('report_date')
        
        if not report_date_str:
            report_date = datetime.now().date()
        else:
            report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
        
        print(f"\n=== ГЕНЕРАЦИЯ ОТЧЕТА ПО ФИКСИРОВАННОМУ ШАБЛОНУ ===")
        print(f"Дата: {report_date}")
        
        # Проверяем существование шаблона
        template_path = os.path.join('report_templates', 'Сводный_отчет_шаблон.xlsx')
        if not os.path.exists(template_path):
            return jsonify({'error': f'Шаблон не найден: {template_path}'}), 400
        
        # Импортируем здесь чтобы избежать циклических импортов
        try:
            from reports.template_report_fixed import FixedTemplateReportGenerator
        except ImportError as e:
            print(f"Ошибка импорта FixedTemplateReportGenerator: {e}")
            return jsonify({'error': f'Не удалось импортировать генератор отчетов: {str(e)}'}), 500
        
        # Создаем генератор
        generator = FixedTemplateReportGenerator(db, template_path)
        
        # Генерируем отчет
        report_path = generator.generate_report(report_date)
        report_filename = os.path.basename(report_path)
        
        # Для GET запросов возвращаем HTML
        if request.method == 'GET':
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Отчет по шаблону</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; }}
                    .success {{ color: green; font-size: 18px; }}
                    .info {{ margin: 20px 0; padding: 15px; background: #f0f8ff; border-radius: 5px; }}
                    .btn {{
                        display: inline-block;
                        padding: 12px 24px;
                        background: #2196F3;
                        color: white;
                        text-decoration: none;
                        border-radius: 5px;
                        margin: 10px 0;
                        font-weight: bold;
                        font-size: 16px;
                    }}
                    .btn:hover {{ background: #1976D2; }}
                    .btn-download {{
                        background: #4CAF50;
                        font-size: 18px;
                        padding: 15px 30px;
                    }}
                    .btn-download:hover {{ background: #45a049; }}
                </style>
            </head>
            <body>
                <h1>📊 Отчет по шаблону успешно сгенерирован!</h1>
                
                <div class="info">
                    <p><strong>📄 Файл:</strong> {report_filename}</p>
                    <p><strong>📅 Дата отчета:</strong> {report_date.strftime('%d.%m.%Y')}</p>
                    <p><strong>🏭 Тип отчета:</strong> Сводный отчет по шаблону Excel</p>
                    <p><em>Отчет содержит заполненные данные из базы данных</em></p>
                </div>
                
                <a class="btn btn-download" href="/download-report/{report_filename}">
                    ⬇️ Скачать отчет Excel
                </a>
                
                <br><br>
                <div style="margin-top: 30px;">
                    <a class="btn" href="/">← На главную</a> | 
                    <a class="btn" href="/generate-template-report">Простой отчет по шаблону</a> | 
                    <a class="btn" href="/generate-report">Стандартный отчет</a>
                </div>
            </body>
            </html>
            """
            return html
        else:
            # Для POST запросов возвращаем JSON
            return jsonify({
                'success': True,
                'report_path': report_path,
                'filename': report_filename,
                'download_url': f'/download-report/{report_filename}',
                'message': 'Отчет по шаблону успешно сгенерирован'
            })
        
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"Ошибка генерации отчета по шаблону: {error_details}")
        
        if request.method == 'GET':
            html = f"""
            <!DOCTYPE html>
            <html>
            <body>
                <h1>❌ Ошибка генерации отчета</h1>
                <p style="color: red;">{str(e)}</p>
                <details>
                    <summary>Подробности ошибки</summary>
                    <pre style="background: #f5f5f5; padding: 10px; overflow: auto;">{error_details}</pre>
                </details>
                <br>
                <a href="/">← На главную</a>
            </body>
            </html>
            """
            return html, 500
        else:
            return jsonify({'error': str(e), 'details': error_details}), 500

@app.route('/generate-from-existing')
def generate_from_existing():
    """Генерация отчета из существующих данных в базе"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        print("\n=== ГЕНЕРАЦИЯ ОТЧЕТА ИЗ СУЩЕСТВУЮЩИХ ДАННЫХ ===")
        
        # Создаем книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводный отчет"
        
        # Заголовок
        ws['A1'] = "СВОДНЫЙ ОТЧЕТ ПО ТОПЛИВООБЕСПЕЧЕНИЮ"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = f"Дата генерации: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws.merge_cells('A2:G2')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws['A3'] = "Данные из базы данных"
        ws.merge_cells('A3:G3')
        ws['A3'].alignment = Alignment(horizontal='center')
        
        # Получаем данные из базы
        session = db_connection.get_session()
        
        # Заголовки таблицы
        headers = ["№", "Компания", "Остатки АИ-92 (т)", "Остатки АИ-95 (т)", 
                  "Поставки АИ-92 (т)", "Реализация АИ-92 (т/мес)", "Дата данных"]
        
        row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        row += 1
        
        # Собираем данные для каждой компании
        companies_data = {}
        
        # Получаем все данные из Sheet3Balance (остатки)
        balances = session.query(Sheet3Balance).all()
        for balance in balances:
            company_name = balance.company_name
            if company_name not in companies_data:
                companies_data[company_name] = {
                    'stock_ai92': 0,
                    'stock_ai95': 0,
                    'supply_ai92': 0,
                    'sales_ai92': 0,
                    'data_date': balance.report_date
                }
            
            companies_data[company_name]['stock_ai92'] += (balance.stock_ai92 or 0)
            companies_data[company_name]['stock_ai95'] += (balance.stock_ai95 or 0)
        
        # Получаем данные из Sheet5Sales (реализация)
        sales = session.query(Sheet5Sales).all()
        for sale in sales:
            company_name = sale.company_name
            if company_name not in companies_data:
                companies_data[company_name] = {
                    'stock_ai92': 0,
                    'stock_ai95': 0,
                    'supply_ai92': 0,
                    'sales_ai92': 0,
                    'data_date': sale.report_date
                }
            
            companies_data[company_name]['sales_ai92'] += (sale.monthly_ai92 or 0)
        
        # Получаем данные из Sheet4Supply (поставки)
        supplies = session.query(Sheet4Supply).all()
        for supply in supplies:
            company_name = supply.company_name
            if company_name not in companies_data:
                companies_data[company_name] = {
                    'stock_ai92': 0,
                    'stock_ai95': 0,
                    'supply_ai92': 0,
                    'sales_ai92': 0,
                    'data_date': supply.report_date
                }
            
            companies_data[company_name]['supply_ai92'] += (supply.supply_ai92 or 0)
        
        # Заполняем таблицу
        idx = 1
        total_stock_ai92 = 0
        total_stock_ai95 = 0
        total_supply_ai92 = 0
        total_sales_ai92 = 0
        
        for company_name, data in companies_data.items():
            if data['stock_ai92'] > 0 or data['stock_ai95'] > 0 or data['sales_ai92'] > 0:
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=company_name)
                ws.cell(row=row, column=3, value=float(data['stock_ai92']))
                ws.cell(row=row, column=4, value=float(data['stock_ai95']))
                ws.cell(row=row, column=5, value=float(data['supply_ai92']))
                ws.cell(row=row, column=6, value=float(data['sales_ai92']))
                ws.cell(row=row, column=7, value=data['data_date'].strftime('%d.%m.%Y') if data['data_date'] else '')
                
                # Форматирование числовых ячеек
                for col in [3, 4, 5, 6]:
                    ws.cell(row=row, column=col).number_format = '0.000'
                
                total_stock_ai92 += data['stock_ai92']
                total_stock_ai95 += data['stock_ai95']
                total_supply_ai92 += data['supply_ai92']
                total_sales_ai92 += data['sales_ai92']
                
                idx += 1
                row += 1
        
        db_connection.close_session()
        
        if idx == 1:
            ws.cell(row=row, column=1, value="Нет данных в базе данных")
            ws.merge_cells(f'A{row}:G{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        else:
            # Итоговая строка
            ws.cell(row=row, column=1, value="ИТОГО:")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"Всего компаний: {idx-1}")
            ws.cell(row=row, column=2).font = Font(bold=True)
            
            ws.cell(row=row, column=3, value=float(total_stock_ai92))
            ws.cell(row=row, column=3).font = Font(bold=True)
            ws.cell(row=row, column=3).number_format = '0.000'
            
            ws.cell(row=row, column=4, value=float(total_stock_ai95))
            ws.cell(row=row, column=4).font = Font(bold=True)
            ws.cell(row=row, column=4).number_format = '0.000'
            
            ws.cell(row=row, column=5, value=float(total_supply_ai92))
            ws.cell(row=row, column=5).font = Font(bold=True)
            ws.cell(row=row, column=5).number_format = '0.000'
            
            ws.cell(row=row, column=6, value=float(total_sales_ai92))
            ws.cell(row=row, column=6).font = Font(bold=True)
            ws.cell(row=row, column=6).number_format = '0.000'
        
        # Автоподбор ширины
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Границы для таблицы
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        for r in range(5, row+1):
            for c in range(1, len(headers)+1):
                ws.cell(row=r, column=c).border = thin_border
        
        # Сохраняем файл
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = os.path.join('reports_output', f'Сводный_отчет_из_БД_{timestamp}.xlsx')
        
        wb.save(output_path)
        
        print(f"Отчет сохранен: {output_path}")
        print(f"Компаний в отчете: {idx-1}")
        print(f"Остатки АИ-92 всего: {total_stock_ai92:.3f} т")
        print(f"Остатки АИ-95 всего: {total_stock_ai95:.3f} т")
        
        # Возвращаем HTML с информацией
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Отчет сгенерирован</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                .success {{ color: green; font-size: 18px; }}
                .info {{ margin: 20px 0; padding: 15px; background: #f0f8ff; border-radius: 5px; }}
                .btn {{
                    display: inline-block;
                    padding: 12px 24px;
                    background: #2196F3;
                    color: white;
                    text-decoration: none;
                    border-radius: 5px;
                    margin: 10px 0;
                    font-weight: bold;
                    font-size: 16px;
                }}
                .btn:hover {{ background: #1976D2; }}
                .btn-download {{
                    background: #4CAF50;
                    font-size: 18px;
                    padding: 15px 30px;
                }}
                .btn-download:hover {{ background: #45a049; }}
                .company-list {{ max-height: 300px; overflow-y: auto; border: 1px solid #ddd; padding: 10px; margin: 10px 0; }}
            </style>
        </head>
        <body>
            <h1>✅ Отчет успешно сгенерирован!</h1>
            
            <div class="info">
                <p><strong>📄 Файл:</strong> Сводный_отчет_из_БД_{timestamp}.xlsx</p>
                <p><strong>🏢 Компаний в отчете:</strong> {idx-1}</p>
                <p><strong>📅 Дата генерации:</strong> {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
                <p><strong>📊 Итоговые данные:</strong></p>
                <ul>
                    <li>Остатки АИ-92 всего: {total_stock_ai92:.3f} т</li>
                    <li>Остатки АИ-95 всего: {total_stock_ai95:.3f} т</li>
                    <li>Поставки АИ-92 всего: {total_supply_ai92:.3f} т</li>
                    <li>Реализация АИ-92 всего: {total_sales_ai92:.3f} т/мес</li>
                </ul>
            </div>
            
            <a class="btn btn-download" href="/download-report/Сводный_отчет_из_БД_{timestamp}.xlsx">
                📥 Скачать отчет Excel
            </a>
            
            <div class="company-list">
                <h3>Компании в отчете:</h3>
                <ul>
        """
        
        for company_name in companies_data.keys():
            data = companies_data[company_name]
            html += f"""
                    <li>
                        <strong>{company_name}</strong>
                        <br>Остатки: АИ-92={data['stock_ai92']:.3f} т, АИ-95={data['stock_ai95']:.3f} т
                        <br>Реализация: АИ-92={data['sales_ai92']:.3f} т/мес
                    </li>
            """
        
        html += """
                </ul>
            </div>
            
            <br><br>
            <div style="margin-top: 30px;">
                <a class="btn" href="/">← На главную</a> | 
                <a class="btn" href="/generate-report">Стандартный отчет</a> | 
                <a class="btn" href="/check-db-data">Проверить данные</a>
            </div>
        </body>
        </html>
        """
        
        return html
        
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"Ошибка генерации отчета: {error_details}")
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <body>
            <h1>❌ Ошибка генерации отчета</h1>
            <p style="color: red;">{str(e)}</p>
            <pre style="background: #f5f5f5; padding: 10px; overflow: auto;">{error_details}</pre>
            <a href="/">← На главную</a>
        </body>
        </html>
        """
        return html

@app.route('/api/recent-files')
def api_recent_files():
    """API для получения последних файлов"""
    try:
        files = db.get_recent_files(limit=20)  # Увеличили лимит
        return jsonify(files)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/companies')
def api_companies():
    """API для получения списка компаний"""
    companies = db.get_companies()
    return jsonify([{
        'id': c.id,
        'name': c.name,
        'code': c.code
    } for c in companies])

@app.route('/test-generate')
def test_generate():
    """Тест генерации отчета без фильтра по дате"""
    try:
        generator = SummaryReportGenerator(db)
        
        # Генерируем отчет по последним данным без фильтра по дате
        report_date = datetime.now().date()
        
        # Получаем данные без фильтра по дате
        session = db_connection.get_session()
        
        # Простой запрос всех компаний с данными
        companies_with_data = {}
        
        # Ищем все компании из БД
        db_companies = session.query(Company).all()
        
        for company in db_companies:
            print(f"\nПроверяем компанию: {company.name} (ID: {company.id})")
            
            # Проверяем есть ли данные в sheet1
            sheet1_count = session.query(Sheet1Structure).filter(
                Sheet1Structure.company_id == company.id
            ).count()
            
            sheet3_count = session.query(Sheet3Balance).filter(
                Sheet3Balance.company_id == company.id
            ).count()
            
            sheet5_count = session.query(Sheet5Sales).filter(
                Sheet5Sales.company_id == company.id
            ).count()
            
            if sheet1_count > 0 or sheet3_count > 0 or sheet5_count > 0:
                print(f"  ✓ Есть данные: Sheet1={sheet1_count}, Sheet3={sheet3_count}, Sheet5={sheet5_count}")
                
                # Собираем простые данные
                company_data = {
                    'name': company.name,
                    'sheet1': [],
                    'sheet3_totals': {},
                    'sheet5_totals': {}
                }
                
                # Данные sheet1
                sheet1_items = session.query(Sheet1Structure).filter(
                    Sheet1Structure.company_id == company.id
                ).all()
                
                for item in sheet1_items:
                    company_data['sheet1'].append({
                        'company_name': item.company_name,
                        'azs_count': item.azs_count or 0
                    })
                
                # Данные sheet3
                sheet3_items = session.query(Sheet3Balance).filter(
                    Sheet3Balance.company_id == company.id
                ).all()
                
                if sheet3_items:
                    company_data['sheet3_totals'] = {
                        'total_stock_ai92': sum(item.stock_ai92 or 0 for item in sheet3_items),
                        'total_stock_ai95': sum(item.stock_ai95 or 0 for item in sheet3_items)
                    }
                
                # Данные sheet5
                sheet5_items = session.query(Sheet5Sales).filter(
                    Sheet5Sales.company_id == company.id
                ).all()
                
                if sheet5_items:
                    company_data['sheet5_totals'] = {
                        'total_monthly_ai92': sum(item.monthly_ai92 or 0 for item in sheet5_items),
                        'total_monthly_ai95': sum(item.monthly_ai95 or 0 for item in sheet5_items)
                    }
                
                companies_with_data[company.name] = company_data
            else:
                print(f"  ✗ Нет данных для компании")
        
        db_connection.close_session()
        
        print(f"\nНайдено компаний с данными: {len(companies_with_data)}")
        
        if companies_with_data:
            # Создаем простой отчет вручную
            from openpyxl import Workbook
            from openpyxl.styles import Font
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Тестовый отчет"
            
            ws['A1'] = "ТЕСТОВЫЙ ОТЧЕТ ПО ДАННЫМ ИЗ БД"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:E1')
            
            ws['A2'] = f"Сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ws.merge_cells('A2:E2')
            
            headers = ["Компания", "АЗС (шт)", "Остатки АИ-92 (т)", "Остатки АИ-95 (т)", "Продажи АИ-92 (т/мес)"]
            row = 4
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
            
            row += 1
            
            total_azs = 0
            total_stock_ai92 = 0
            total_stock_ai95 = 0
            total_sales_ai92 = 0
            
            for company_name, data in companies_with_data.items():
                # Считаем общее количество АЗС
                azs_count = sum(item.get('azs_count', 0) for item in data.get('sheet1', []))
                stock_ai92 = data.get('sheet3_totals', {}).get('total_stock_ai92', 0)
                stock_ai95 = data.get('sheet3_totals', {}).get('total_stock_ai95', 0)
                sales_ai92 = data.get('sheet5_totals', {}).get('total_monthly_ai92', 0)
                
                ws.cell(row=row, column=1, value=company_name)
                ws.cell(row=row, column=2, value=azs_count)
                ws.cell(row=row, column=3, value=stock_ai92)
                ws.cell(row=row, column=4, value=stock_ai95)
                ws.cell(row=row, column=5, value=sales_ai92)
                
                total_azs += azs_count
                total_stock_ai92 += stock_ai92
                total_stock_ai95 += stock_ai95
                total_sales_ai92 += sales_ai92
                
                row += 1
            
            # Итоговая строка
            ws.cell(row=row, column=1, value="ИТОГО:")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=total_azs)
            ws.cell(row=row, column=3, value=total_stock_ai92)
            ws.cell(row=row, column=4, value=total_stock_ai95)
            ws.cell(row=row, column=5, value=total_sales_ai92)
            
            # Форматирование числовых ячеек
            for r in range(5, row+1):
                for c in range(2, 6):
                    cell = ws.cell(row=r, column=c)
                    cell.number_format = '0.00'
            
            # Автоподбор ширины
            from openpyxl.utils import get_column_letter
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = os.path.join('reports_output', f'Тестовый_отчет_{timestamp}.xlsx')
            
            wb.save(output_path)
            
            return jsonify({
                'success': True,
                'message': f'Тестовый отчет создан с {len(companies_with_data)} компаниями',
                'path': output_path,
                'filename': os.path.basename(output_path),
                'companies_count': len(companies_with_data),
                'total_azs': total_azs,
                'total_stock_ai92': round(total_stock_ai92, 2),
                'total_stock_ai95': round(total_stock_ai95, 2),
                'total_sales_ai92': round(total_sales_ai92, 2)
            })
        else:
            return jsonify({'error': 'Нет данных в БД'}), 400
            
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"Ошибка тестовой генерации: {error_details}")
        return jsonify({'error': str(e), 'details': error_details}), 500

@app.route('/debug-data')
def debug_data():
    """Отладочная информация о данных"""
    try:
        # Получаем сегодняшнюю дату
        today = datetime.now().date()
        
        print(f"\n=== ОТЛАДКА ДАННЫХ ====")
        print(f"Запрашиваем данные на дату: {today}")
        
        # Получаем агрегированные данные БЕЗ фильтра по дате
        print("\n1. Получаем данные БЕЗ фильтра по дате:")
        aggregated_no_date = db.get_aggregated_data()  # Без аргументов
        
        print(f"   Найдено компаний: {len(aggregated_no_date)}")
        for name, data in aggregated_no_date.items():
            print(f"   - {name}: sheet1={len(data.get('sheet1', []))}, sheet3={len(data.get('sheet3_data', []))}, sheet5={len(data.get('sheet5_data', []))}")
        
        # Получаем агрегированные данные С фильтром по дате
        print("\n2. Получаем данные С фильтром по дате (сегодня):")
        aggregated_today = db.get_aggregated_data(today)
        
        print(f"   Найдено компаний: {len(aggregated_today)}")
        for name, data in aggregated_today.items():
            print(f"   - {name}: sheet1={len(data.get('sheet1', []))}, sheet3={len(data.get('sheet3_data', []))}, sheet5={len(data.get('sheet5_data', []))}")
        
        # Проверяем какие даты есть в БД
        print("\n3. Проверяем даты в БД:")
        session = db_connection.get_session()
        
        # Даты из UploadedFile
        file_dates = session.query(UploadedFile.report_date).distinct().all()
        print(f"   Даты файлов: {[str(d[0]) for d in file_dates]}")
        
        # Даты из Sheet1Structure
        sheet1_dates = session.query(Sheet1Structure.report_date).distinct().all()
        print(f"   Даты Sheet1: {[str(d[0]) for d in sheet1_dates]}")
        
        # Даты из Sheet5Sales
        sheet5_dates = session.query(Sheet5Sales.report_date).distinct().all()
        print(f"   Даты Sheet5: {[str(d[0]) for d in sheet5_dates]}")
        
        db_connection.close_session()
        
        return jsonify({
            'success': True,
            'without_date_filter': len(aggregated_no_date),
            'with_date_filter': len(aggregated_today),
            'file_dates': [str(d[0]) for d in file_dates],
            'sheet1_dates': [str(d[0]) for d in sheet1_dates],
            'sheet5_dates': [str(d[0]) for d in sheet5_dates],
            'message': f'Без фильтра: {len(aggregated_no_date)} компаний, С фильтром: {len(aggregated_today)} компаний'
        })
        
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"Ошибка отладки: {error_details}")
        return jsonify({'error': str(e), 'details': error_details}), 500

# @app.route('/generate-template-report', methods=['GET', 'POST'])
# def generate_template_report():
#     """Генерация отчета по шаблону"""
#     try:
#         # Получаем дату отчета
#         if request.method == 'POST':
#             if request.is_json:
#                 data = request.get_json()
#                 report_date_str = data.get('report_date')
#             else:
#                 report_date_str = request.form.get('report_date')
#         else:
#             report_date_str = request.args.get('report_date')
        
#         if not report_date_str:
#             report_date = datetime.now().date()
#         else:
#             report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
        
#         print(f"\n=== ГЕНЕРАЦИЯ ОТЧЕТА ПО ШАБЛОНУ ===")
#         print(f"Дата: {report_date}")
        
#         # Проверяем существование папки для шаблонов
#         template_dir = 'report_templates'
#         if not os.path.exists(template_dir):
#             os.makedirs(template_dir)
#             print(f"Создана папка для шаблонов: {template_dir}")
        
#         # Ищем шаблон
#         template_path = os.path.join(template_dir, 'Сводный_отчет_шаблон.xlsx')
        
#         if not os.path.exists(template_path):
#             # Создаем простой шаблон если его нет
#             return create_sample_template_response(template_path)
        
#         # Создаем генератор
#         template_generator = TemplateReportGenerator(db, template_path)
        
#         # Генерируем отчет
#         report_path = template_generator.generate_from_template(report_date)
#         report_filename = os.path.basename(report_path)
        
#         # Для GET запросов возвращаем HTML
#         if request.method == 'GET':
#             return render_template_report_success(report_filename, report_date)
#         else:
#             # Для POST запросов возвращаем JSON
#             return jsonify({
#                 'success': True,
#                 'report_path': report_path,
#                 'filename': report_filename,
#                 'download_url': f'/download-report/{report_filename}',
#                 'message': 'Отчет по шаблону успешно сгенерирован'
#             })
        
#     except Exception as e:
#         error_details = traceback.format_exc()
#         print(f"Ошибка генерации отчета по шаблону: {error_details}")
        
#         if request.method == 'GET':
#             return render_template_error(str(e), error_details)
#         else:
#             return jsonify({'error': str(e), 'details': error_details}), 500

def render_template_report_success(filename, report_date):
    """Рендеринг страницы успешной генерации"""
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Отчет по шаблону</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; }}
            .success {{ color: green; font-size: 18px; }}
            .info {{ margin: 20px 0; padding: 15px; background: #f0f8ff; border-radius: 5px; }}
            .btn {{
                display: inline-block;
                padding: 12px 24px;
                background: #2196F3;
                color: white;
                text-decoration: none;
                border-radius: 5px;
                margin: 10px 0;
                font-weight: bold;
                font-size: 16px;
            }}
            .btn:hover {{ background: #1976D2; }}
            .btn-download {{
                background: #4CAF50;
                font-size: 18px;
                padding: 15px 30px;
            }}
            .btn-download:hover {{ background: #45a049; }}
        </style>
    </head>
    <body>
        <h1>📊 Отчет по шаблону успешно сгенерирован!</h1>
        
        <div class="info">
            <p><strong>📄 Файл:</strong> {filename}</p>
            <p><strong>📅 Дата отчета:</strong> {report_date.strftime('%d.%m.%Y')}</p>
            <p><strong>🏭 Тип отчета:</strong> Сводный отчет по шаблону Excel</p>
            <p><em>Отчет содержит: структуру, потребность, остатки и реализацию топлива</em></p>
        </div>
        
        <a class="btn btn-download" href="/download-report/{filename}">
            ⬇️ Скачать отчет Excel
        </a>
        
        <br><br>
        <div style="margin-top: 30px;">
            <a class="btn" href="/">← На главную</a> | 
            <a class="btn" href="/generate-report">Стандартный отчет</a> | 
            <a class="btn" href="/test-generate">Тестовая генерация</a>
        </div>
    </body>
    </html>
    """
    return html

def render_template_error(error_msg, error_details):
    """Рендеринг страницы ошибки"""
    html = f"""
    <!DOCTYPE html>
    <html>
    <body>
        <h1>❌ Ошибка генерации отчета</h1>
        <p style="color: red;">{error_msg}</p>
        <p>Убедитесь что:</p>
        <ul>
            <li>Файл шаблона "Сводный_отчет_шаблон.xlsx" находится в папке report_templates</li>
            <li>Шаблон имеет правильные имена листов</li>
            <li>В базе данных есть загруженные файлы</li>
        </ul>
        <details>
            <summary>Подробности ошибки</summary>
            <pre style="background: #f5f5f5; padding: 10px; overflow: auto;">{error_details}</pre>
        </details>
        <br>
        <a href="/">← На главную</a>
    </body>
    </html>
    """
    return html

def create_sample_template_response(template_path):
    """Создание образца шаблона"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        # Создаем простой шаблон
        wb = Workbook()
        
        # Лист 1: Титульный
        ws1 = wb.active
        ws1.title = "1-Титульный"
        ws1['A1'] = "СВОДНЫЙ ОТЧЕТ ПО ТОПЛИВООБЕСПЕЧЕНИЮ"
        ws1['A1'].font = Font(size=16, bold=True)
        ws1['A3'] = "Дата отчета: {дата_отчета}"
        ws1['A4'] = "Количество компаний: {кол_во_компаний}"
        ws1['A5'] = "Всего АЗС: {всего_азс}"
        
        # Лист 2: Структура
        ws2 = wb.create_sheet(title="2-Структура")
        headers = ["№", "Компания", "АЗС", "Работающих АЗС"]
        for col, header in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        # Лист 3: Потребность
        ws3 = wb.create_sheet(title="3-Потребность")
        headers = ["№", "Компания", "Бензин всего", "Дизель всего", "Бензин (мес)", "Дизель (мес)"]
        for col, header in enumerate(headers, 1):
            ws3.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        # Лист 4: Остатки
        ws4 = wb.create_sheet(title="4-Остатки")
        headers = ["№", "Компания", "АИ-92", "АИ-95", "Дизель зим.", "Дизель аркт."]
        for col, header in enumerate(headers, 1):
            ws4.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        # Лист 5: Реализация
        ws5 = wb.create_sheet(title="5-Реализация")
        headers = ["№", "Компания", "АИ-92", "АИ-95", "Дизель зим.", "Дизель аркт."]
        for col, header in enumerate(headers, 1):
            ws5.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        # Сохраняем
        wb.save(template_path)
        print(f"Создан образец шаблона: {template_path}")
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <body>
            <h1>📋 Создан образец шаблона</h1>
            <p>Файл шаблона не был найден, поэтому создан образец: <strong>Сводный_отчет_шаблон.xlsx</strong></p>
            <p>Пожалуйста:</p>
            <ol>
                <li>Откройте файл: {template_path}</li>
                <li>Настройте форматирование по вашему вкусу</li>
                <li>Сохраните файл</li>
                <li><a href="/generate-template-report">Попробуйте снова</a></li>
            </ol>
            <a href="/">← На главную</a>
        </body>
        </html>
        """
        return html
        
    except Exception as e:
        return f"Ошибка создания шаблона: {str(e)}"

@app.route('/test-parse')
def test_parse():
    """Тестовая страница для проверки парсинга"""
    try:
        # Ищем любой Excel файл в папке uploads
        uploads_dir = app.config['UPLOAD_FOLDER']
        excel_files = [f for f in os.listdir(uploads_dir) if f.endswith('.xlsx')]
        
        if not excel_files:
            return "Нет Excel файлов для тестирования. Сначала загрузите файл через форму."
        
        test_file = os.path.join(uploads_dir, excel_files[0])
        
        result_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Тест парсинга</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                table {{ border-collapse: collapse; margin: 10px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .success {{ color: green; }}
                .error {{ color: red; }}
                .warning {{ color: orange; }}
                pre {{ background-color: #f5f5f5; padding: 10px; overflow-x: auto; }}
                .parser-section {{ border: 1px solid #ccc; padding: 15px; margin: 10px 0; border-radius: 5px; }}
            </style>
        </head>
        <body>
            <h1>Тест парсинга файла: {excel_files[0]}</h1>
        """
        
        # Пробуем новый улучшенный парсер
        if NEW_PARSER_AVAILABLE:
            try:
                parser = SimpleAllParserV2(test_file)
                all_data = parser.parse_all()
                metadata = all_data['metadata']
                
                result_html += f"""
                <div class="parser-section">
                    <h2 class="success">✅ Новый улучшенный парсер работает</h2>
                    <p><strong>Компания:</strong> {metadata['company']}</p>
                    <p><strong>Дата отчета:</strong> {metadata['report_date']}</p>
                    <p><strong>Исполнитель:</strong> {metadata.get('executor', 'не указан')}</p>
                    
                    <h3>Извлеченные данные:</h3>
                    <ul>
                        <li>Лист 1 (Структура): {len(all_data['sheet1'])} записей</li>
                        <li>Лист 2 (Потребность): {'данные есть' if all_data['sheet2'] else 'нет данных'}</li>
                        <li>Лист 3 (Остатки): {len(all_data['sheet3'])} записей</li>
                        <li>Лист 4 (Поставка): {len(all_data['sheet4'])} записей</li>
                        <li>Лист 5 (Реализация): {len(all_data['sheet5'])} записей</li>
                        <li>Лист 6 (Авиатопливо): {len(all_data['sheet6'])} записей</li>
                        <li>Лист 7 (Справка): {len(all_data['sheet7'])} записей</li>
                    </ul>
                """
                
                # Показываем пример данных
                if all_data['sheet1']:
                    result_html += """
                    <h4>Пример данных из листа 1 (первые 5 записей):</h4>
                    <table>
                        <tr>
                            <th>Принадлежность</th>
                            <th>Название компании</th>
                            <th>Нефтебаз</th>
                            <th>АЗС</th>
                            <th>Работающих АЗС</th>
                        </tr>
                    """
                    
                    for i, item in enumerate(all_data['sheet1'][:5]):
                        result_html += f"""
                        <tr>
                            <td>{item.get('affiliation', '')[:50]}</td>
                            <td>{item.get('company_name', '')[:50]}</td>
                            <td>{item.get('oil_depots_count', 0)}</td>
                            <td>{item.get('azs_count', 0)}</td>
                            <td>{item.get('working_azs_count', 0)}</td>
                        </tr>
                        """
                    
                    result_html += "</table>"
                
                result_html += "</div>"
                
            except Exception as e:
                result_html += f"""
                <div class="parser-section">
                    <h2 class="error">❌ Новый улучшенный парсер не работает: {str(e)}</h2>
                    <pre>{traceback.format_exc()}</pre>
                </div>
                """
        
        # Пробуем упрощенный парсер всех листов
        if SIMPLE_ALL_PARSER_AVAILABLE:
            try:
                parser = SimpleAllParser(test_file)
                data = parser.parse_all()
                
                result_html += f"""
                <div class="parser-section">
                    <h2 class="success">✅ Упрощенный парсер всех листов работает</h2>
                    <p><strong>Компания:</strong> {data.get('metadata', {}).get('company', 'Неизвестно')}</p>
                    
                    <h3>Извлеченные данные:</h3>
                    <ul>
                        <li>Лист 1 (Структура): {len(data.get('sheet1', []))} записей</li>
                        <li>Лист 2 (Потребность): {'данные есть' if data.get('sheet2') else 'нет данных'}</li>
                        <li>Лист 3 (Остатки): {len(data.get('sheet3', []))} записей</li>
                        <li>Лист 4 (Поставка): {len(data.get('sheet4', []))} записей</li>
                        <li>Лист 5 (Реализация): {len(data.get('sheet5', []))} записей</li>
                    </ul>
                </div>
                """
                
            except Exception as e:
                result_html += f"""
                <div class="parser-section">
                    <h2 class="error">❌ Упрощенный парсер всех листов не работает: {str(e)}</h2>
                    <pre>{traceback.format_exc()}</pre>
                </div>
                """
        
        # Пробуем основной парсер
        if PARSER_AVAILABLE:
            try:
                parser = FuelReportParser(test_file)
                metadata = parser.parse()
                
                result_html += f"""
                <div class="parser-section">
                    <h2 class="success">✅ Основной парсер работает</h2>
                    <p><strong>Компания:</strong> {metadata.company_name}</p>
                    <p><strong>Дата отчета:</strong> {metadata.report_date}</p>
                """
                
                try:
                    all_data = parser.extract_all_data()
                    result_html += f"""
                    <h3>Извлеченные данные:</h3>
                    <ul>
                        <li>Лист 1 (Структура): {len(all_data.get('sheet1', []))} записей</li>
                        <li>Лист 2 (Потребность): {'данные есть' if all_data.get('sheet2') else 'нет данных'}</li>
                        <li>Лист 3 (Остатки): {len(all_data.get('sheet3', []))} записей</li>
                        <li>Лист 4 (Поставка): {len(all_data.get('sheet4', []))} записей</li>
                        <li>Лист 5 (Реализация): {len(all_data.get('sheet5', []))} записей</li>
                    </ul>
                    """
                except Exception as e:
                    result_html += f"""
                    <h3 class="error">❌ Ошибка при извлечении данных: {str(e)}</h3>
                    """
                
                result_html += "</div>"
                
            except Exception as e:
                result_html += f"""
                <div class="parser-section">
                    <h2 class="error">❌ Основной парсер не работает: {str(e)}</h2>
                    <pre>{traceback.format_exc()}</pre>
                </div>
                """
        
        # Пробуем простой парсер
        if SIMPLE_PARSER_AVAILABLE:
            try:
                parser = SimpleFuelParser(test_file)
                data = parser.parse_all()
                
                result_html += f"""
                <div class="parser-section">
                    <h2 class="success">✅ Простой парсер работает</h2>
                    <p><strong>Компания:</strong> {data.get('company', 'Неизвестно')}</p>
                    
                    <h3>Извлеченные данные:</h3>
                    <ul>
                        <li>Лист 1 (Структура): {len(data.get('sheet1', []))} записей</li>
                        <li>Лист 2 (Потребность): {'данные есть' if data.get('sheet2') else 'нет данных'}</li>
                        <li>Лист 3 (Остатки): {len(data.get('sheet3', []))} записей</li>
                    </ul>
                </div>
                """
                
            except Exception as e:
                result_html += f"""
                <div class="parser-section">
                    <h2 class="error">❌ Простой парсер не работает: {str(e)}</h2>
                    <pre>{traceback.format_exc()}</pre>
                </div>
                """
        
        result_html += """
            <hr>
            <p><a href="/">Вернуться на главную</a></p>
        </body>
        </html>
        """
        
        return result_html
        
    except Exception as e:
        return f"Ошибка: {str(e)}<br><pre>{traceback.format_exc()}</pre>"

if __name__ == '__main__':
    print("=" * 50)
    print("Система отчетов по топливообеспечению")
    print(f"Парсеры доступны:")
    print(f"  Новый улучшенный: {NEW_PARSER_AVAILABLE}")
    print(f"  Основной: {PARSER_AVAILABLE}")
    print(f"  Упрощенный всех листов: {SIMPLE_ALL_PARSER_AVAILABLE}")
    print(f"  Простой: {SIMPLE_PARSER_AVAILABLE}")
    print("=" * 50)
    print("Доступные эндпоинты:")
    print("  GET  /              - Главная страница")
    print("  POST /upload        - Загрузка файла (новый парсер)")
    print("  POST /generate-report - Генерация отчета")
    print("  GET  /download-report/<filename> - Скачивание отчета")
    print("  GET  /test-parse    - Тест парсинга файла")
    print("  GET  /api/recent-files - API последних файлов")
    print("  GET  /api/companies - API списка компаний")
    print("=" * 50)
    
    # Создаем необходимые папки
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(app.config['REPORTS_FOLDER'], exist_ok=True)
    os.makedirs('reports_output', exist_ok=True)
    
    app.run(debug=True, host='0.0.0.0', port=5000)